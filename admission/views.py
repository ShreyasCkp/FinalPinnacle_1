from pyexpat import model
from django.shortcuts import render, redirect
from .forms import PUAdmissionForm, DegreeAdmissionForm
from .models import PUAdmission, DegreeAdmission
from .models import Enquiry1
from master.decorators import custom_login_required
# ---------- PU Admission Form View (AUTOINCREMENT FIXED) ----------

import datetime
from django.shortcuts import render


@custom_login_required
def admission_list(request):
    admissions = PUAdmission.objects.all()
    admissions = sorted(admissions, key=lambda a: int(a.admission_no.split('/')[-1]))

    # Extract all user IDs where admission_taken_by is not None
    user_ids = [a.admission_taken_by for a in admissions if a.admission_taken_by]

    # Fetch corresponding users with their usernames
    users = UserCustom.objects.filter(id__in=user_ids).values('id', 'username')

    # Create a dictionary to map user IDs to usernames
    user_dict = {user['id']: user['username'] for user in users}

    # Add the username dynamically to each admission object
    for admission in admissions:
        admission.created_by_username = user_dict.get(admission.admission_taken_by, "Unknown")

         # Label based on admission_source saved in DB
        if admission.admission_source == 'enquiry':
            admission.taken_from_label = "From Enquiry"
        elif admission.admission_source == 'bulk_import':
            admission.taken_from_label = "From Bulk"
        else:
            admission.taken_from_label = "Direct"

    return render(request, 'admission/admission_list.html', {'admissions': admissions})
 
 
 
 
from django.shortcuts import redirect
from django.http import JsonResponse
from django.contrib import messages
import pandas as pd
import pdfplumber
import re
from datetime import datetime
from admission.models import PUAdmission, DegreeAdmission
from master.models import Course
from django.db import models



@custom_login_required
def parse_date_safe(value):
    try:
        if pd.isna(value) or value in ['', 'NaT', 'None']:
            return None
        if isinstance(value, datetime):
            return value.date()
        return pd.to_datetime(value, errors='coerce').date()
    except Exception:
        return None


    
@custom_login_required
def normalize_category(value):
    valid_categories = dict(PUAdmission.CATEGORY_CHOICES)
    value = str(value).strip().upper()
    return value if value in valid_categories else 'GM'


@custom_login_required
def get_course_object(course_name):
    if not course_name:
        return None
    return Course.objects.filter(name__icontains=str(course_name).strip()).first()


@custom_login_required
def bulk_import_admissions(request):
    if request.method == 'POST' and request.FILES.get('import_file'):
        file = request.FILES['import_file']
        filename = file.name.lower()

        errors = []
        count = 0

        # Fetch current serials for PU and Degree
        pu_prefix = "PSCM2025-26PUC"
        deg_prefix = "PSCM2025-26DG"

        last_pu = PUAdmission.objects.filter(admission_no__startswith=f"{pu_prefix}/").order_by('-id').first()
        pu_serial = int(re.search(r"/(\d+)$", last_pu.admission_no).group(1)) if last_pu and last_pu.admission_no else 0

        last_deg = DegreeAdmission.objects.filter(admission_no__startswith=f"{deg_prefix}/").order_by('-id').first()
        deg_serial = int(re.search(r"/(\d+)$", last_deg.admission_no).group(1)) if last_deg and last_deg.admission_no else 0

        try:
            if filename.endswith('.pdf'):
                print("📄 Reading PDF")
                rows = []
                with pdfplumber.open(file) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text()
                        if not text:
                            continue
                        for line in text.split('\n'):
                            if 'admission_no' in line.lower():
                                continue
                            line = re.sub(r'\s{2,}', ' ', line).strip()
                            parts = line.split()

                            try:
                                mobile = parts[-1] if re.match(r'^\d{10}$', parts[-1]) else ''
                                category = parts[-2] if len(parts) >= 2 else ''
                                gender = parts[-3] if len(parts) >= 3 else ''
                                dob = parts[-4] if len(parts) >= 4 else ''
                                dob_parsed = parse_date_safe(dob)
                                slice_offset = 4 if mobile else 3
                                rest = parts[:-slice_offset]

                                admission_no = rest[0] if rest else ''
                                course = next((t for t in rest if "PUC" in t), None)
                                if not course:
                                    raise ValueError("Course not found")
                                course_idx = rest.index(course)
                                student_name = ' '.join(rest[1:course_idx])
                                parent_name = ' '.join(rest[course_idx + 1:])

                                rows.append({
                                    'admission_no': admission_no,
                                    'student_name': student_name,
                                    'course': course,
                                    'parent_name': parent_name,
                                    'dob': dob_parsed,
                                    'gender': gender,
                                    'categoy': category,
                                    'parent_mobile_no': mobile,
                                })
                            except Exception as e:
                                errors.append(f"⚠️ Line skipped: {line} → {e}")
                df = pd.DataFrame(rows)

            elif filename.endswith(('.xlsx', '.xls')):
                print("📊 Reading Excel")
                df = pd.read_excel(file)
                df = df.fillna('')
            else:
                return JsonResponse({'success': False, 'error': '❌ Unsupported file format'})

            if df.empty:
                return JsonResponse({'success': False, 'error': '⚠️ No data found in file.'})

            for _, row in df.iterrows():
                data = {}
                for field in PUAdmission._meta.get_fields():
                    if field.name in df.columns:
                        value = row.get(field.name, '')

                        if isinstance(field, models.DateField):
                            data[field.name] = parse_date_safe(str(value).strip())
                        elif isinstance(field, models.DecimalField):
                            try:
                                data[field.name] = float(value) if str(value).strip() else None
                            except:
                                data[field.name] = None
                        elif isinstance(field, models.BooleanField):
                            data[field.name] = str(value).strip().lower() in ['true', '1', 'yes']
                        elif isinstance(field, models.ForeignKey):
                            if field.related_model.__name__ == 'Course':
                                data[field.name] = get_course_object(value)
                            else:
                                data[field.name] = None
                        else:
                            data[field.name] = str(value).strip() if pd.notna(value) else ''

                course = data.get('course')
                course_type_name = course.course_type.name.upper() if course and course.course_type else ''

                if not data.get('admission_no'):
                    if 'PU' in course_type_name:
                        data['admission_no'] = f"{pu_prefix}/{pu_serial + 1:03d}"
                        pu_serial += 1
                    elif 'DEGREE' in course_type_name or 'UG' in course_type_name:
                        data['admission_no'] = f"{deg_prefix}/{deg_serial + 1:03d}"
                        deg_serial += 1

                if not data.get('student_name') or not data.get('dob'):
                    errors.append(f"⚠️ Skipped row: Missing student_name or dob → {row.to_dict()}")
                    continue

                try:
                    data['admission_source'] = 'bulk_import'
                    if 'PU' in course_type_name:
                        PUAdmission.objects.create(**data)
                    elif 'DEGREE' in course_type_name or 'UG' in course_type_name:
                        DegreeAdmission.objects.create(**data)
                    else:
                        errors.append(f"⚠️ Unknown course type for: {row.get('course')}")
                        continue

                    print(f"✅ Saved: {data.get('admission_no')} - {data.get('student_name')}")
                    count += 1
                except Exception as e:
                    errors.append(f"❌ {data.get('student_name')} - DB Error: {e}")

            return JsonResponse({
                'success': True,
                'message': f"✅ Successfully imported {count} record(s).",
                'errors': errors[:10],
                'error_count': len(errors)
            })

        except Exception as e:
            print("❌ Import error:", e)
            return JsonResponse({'success': False, 'error': f"❌ Import failed: {str(e)}"})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})

from datetime import date, datetime
from django.shortcuts import render
from .models import PUAdmission, Enquiry1
from .forms import PUAdmissionForm


@custom_login_required
def generate_next_receipt_no():
    today_str = datetime.today().strftime('%Y%m%d')
    latest = PUAdmission.objects.filter(receipt_no__startswith=today_str).order_by('-receipt_no').first()
    if latest and latest.receipt_no:
        try:
            current_inc = int(latest.receipt_no.split('-')[1])
        except (IndexError, ValueError):
            current_inc = 0
        next_inc = current_inc + 1
        last_receipt_no = latest.receipt_no
    else:
        next_inc = 1
        last_receipt_no = None
    next_receipt_no = f"{today_str}-{next_inc:03d}"
    return next_receipt_no, last_receipt_no


from .utils import generate_next_receipt_no_shared
from core.utils import get_logged_in_user, log_activity
import re
from datetime import date
from django.shortcuts import render
from django.http import JsonResponse
from admission.models import PUAdmission, Enquiry1
from master.models import FeeMaster
from admission.forms import PUAdmissionForm



def match_fee_name(name):
    name = name.strip().lower().replace(" ", "")
    if any(prefix in name for prefix in ["tution", "tuition"]):
        return "tuition_fee"
    elif "application" in name:
        return "application_fee"
    elif "books" in name:
        return "books_fee"
    elif "uniform" in name:
        return "uniform_fee"
    elif "transport" in name:
        return "transport_fees"
    return None

@custom_login_required
def get_fees_for_course_type(request):
    course_type_id = request.GET.get("course_type_id")
    course_id = request.GET.get("course_id")

    fees = {
        "tuition_fee": 0,
        "books_fee": 0,
        "uniform_fee": 0,
        "application_fee": 0,
        "transport_fees": 0
    }

    if course_type_id and course_id:
        fee_qs = FeeMaster.objects.filter(program_type_id=course_type_id, combination_id=course_id)
        for fee in fee_qs:
            key = match_fee_name(fee.fee_name)
            if key:
                fees[key] = fee.fee_amount

    return JsonResponse(fees)

from django.db.models import Q
from master.models import CourseType  # 👈 Make sure you import CourseType correctly
from django.shortcuts import render, redirect

@custom_login_required
def admission_form(request, enquiry_no=None):
    success = False
    academic_prefix = "PSCM2025-26PUC"
    last = PUAdmission.objects.filter(admission_no__startswith=f"{academic_prefix}/").order_by('-id').first()
    last_number = int(re.search(r"/(\d+)$", last.admission_no).group(1)) if (last and last.admission_no) else 0

    next_serial = last_number + 1
    next_admission_no = f"{academic_prefix}/{next_serial:03d}"
    next_receipt_no, last_receipt_no = generate_next_receipt_no_shared()

    marks_initial = {}
    for i in range(1, 7):
        marks_initial[f"subject{i}"] = ""
        marks_initial[f"max_marks{i}"] = ""
        marks_initial[f"marks_obtained{i}"] = ""
        marks_initial[f"total_marks_percentage{i}"] = ""

    tuition_fee = books_fee = uniform_fee = application_fee = transport_fees = 0
    course_type_id = course_id = None

    if enquiry_no:
        try:
            enquiry = Enquiry1.objects.get(enquiry_no=enquiry_no)
            course_type_id = enquiry.course_type.id if enquiry.course_type else None
            course_id = enquiry.course.id if enquiry.course else None
        except Enquiry1.DoesNotExist:
            pass

    # 🔍 Identify default PU-type program from CourseType model
    default_course_type = CourseType.objects.filter(
        Q(name__icontains="PU") | Q(name__icontains="PUC")
    ).first()

    form = None

    if request.method == "POST":
        form = PUAdmissionForm(request.POST, request.FILES)
        if form.is_valid():
            ct = form.cleaned_data.get("course_type")
            c = form.cleaned_data.get("course")
            if ct:
                course_type_id = ct.id
            if c:
                course_id = c.id

            if course_type_id and course_id:
                fee_qs = FeeMaster.objects.filter(program_type_id=course_type_id, combination_id=course_id)
                for fee in fee_qs:
                    key = match_fee_name(fee.fee_name)
                    if key:
                        locals()[key] = fee.fee_amount
    else:
        initial_data = {
            **marks_initial,
            'admission_no': next_admission_no,
            'admission_date': date.today(),
            'receipt_no': next_receipt_no,
            'receipt_date': date.today(),
            'application_fee': application_fee,
            'tuition_fee': tuition_fee,
            'books_fee': books_fee,
            'uniform_fee': uniform_fee,
            'transport_fees': transport_fees,
            'course_type': default_course_type.id if default_course_type else None

        }
        if enquiry_no:
            try:
                enquiry = Enquiry1.objects.get(enquiry_no=enquiry_no)
                initial_data.update({
                    'enquiry_no': enquiry.enquiry_no,
                    'student_name': enquiry.student_name,
                    'gender': enquiry.gender,
                    'parent_name': enquiry.parent_name,
                    'parent_mobile_no': enquiry.parent_phone,
                    'email': enquiry.email,
                    'course_type': enquiry.course_type.id if enquiry.course_type else (
                            default_course_type.id if default_course_type else None
                        ),
                    'course': enquiry.course.id if enquiry.course else None,
                    'sslc_percentage': enquiry.percentage_10th,
                })
            except Enquiry1.DoesNotExist:
                pass

        form = PUAdmissionForm(initial=initial_data)

        if course_type_id and course_id:
            fee_qs = FeeMaster.objects.filter(program_type_id=course_type_id, combination_id=course_id)
            for fee in fee_qs:
                key = match_fee_name(fee.fee_name)
                if key:
                    locals()[key] = fee.fee_amount

    if request.method == "POST" and form and form.is_valid():
        pu_admission = form.save(commit=False)
        pu_admission.admission_no = next_admission_no
        pu_admission.receipt_no = next_receipt_no
        pu_admission.receipt_date = date.today()

        student_name = form.cleaned_data.get('student_name')
        enquiry_from_form = form.cleaned_data.get('enquiry_no')

        # 👉 Convert student_name to uppercase (safe even if already in caps)
        if student_name:
            student_name = student_name.upper()

        # 👉 Now match with Enquiry
        if not enquiry_no and not enquiry_from_form:
            enquiry_obj = Enquiry1.objects.filter(student_name=student_name).first()
            enquiry_no = enquiry_obj.enquiry_no if enquiry_obj else None


        pu_admission.enquiry_no = enquiry_no or enquiry_from_form
        pu_admission.admission_taken_by = request.session.get('user_id')

        pu_admission.admission_source = 'enquiry' if enquiry_from_form else 'direct'

        if not pu_admission.admission_date:
            pu_admission.admission_date = date.today()

        pu_admission.application_fee = application_fee
        pu_admission.tuition_fee = tuition_fee
        pu_admission.books_fee = books_fee
        pu_admission.uniform_fee = uniform_fee
        if hasattr(pu_admission, "transport_fees"):
            pu_admission.transport_fees = transport_fees
        elif hasattr(pu_admission, "transport_amount"):
            pu_admission.transport_amount = transport_fees

        tuition_advance = pu_admission.tuition_advance_amount or 0
        scholarship = pu_admission.scholarship_amount or 0
        pu_admission.final_fee_after_advance = tuition_fee - tuition_advance - scholarship

        pu_admission.save()

        user = get_logged_in_user(request)
        log_activity(user, 'created', pu_admission)

        return HttpResponseRedirect(reverse('admission_list') + '?success=add')

        success = True

        snackbar_message = f"Admission {pu_admission.admission_no} saved successfully for {pu_admission.student_name}."

        next_serial += 1
        next_admission_no = f"{academic_prefix}/{next_serial:03d}"
        next_receipt_no, last_receipt_no = generate_next_receipt_no_shared()

        form = PUAdmissionForm(initial={
            **marks_initial,
            'admission_no': next_admission_no,
            'admission_date': date.today(),
            'receipt_no': next_receipt_no,
            'receipt_date': date.today(),
            'application_fee': application_fee,
            'tuition_fee': tuition_fee,
            'books_fee': books_fee,
            'uniform_fee': uniform_fee,
            'transport_fees': transport_fees
        })

    subject_rows = [
        {
            "subject_field": form[f"subject{i}"],
            "max_marks_field": form[f"max_marks{i}"],
            "marks_field": form[f"marks_obtained{i}"],
            "percent_field": form[f"total_marks_percentage{i}"],
            "subject_val": "", "max_marks_val": "", "marks_val": "", "percent_val": ""
        } for i in range(1, 7)
    ]

    return render(request, 'admission/admission_form.html', {
        'form': form,
        'success': success,
        'next_admission_no': next_admission_no,
        'next_receipt_no': next_receipt_no,
        'last_receipt_no': last_receipt_no,
        'subject_rows': subject_rows,
        'add_mode': True,
        'edit_mode': False,
        'view_mode': False,
        'admission': pu_admission if success else None,
        'snackbar_message': snackbar_message if success else "",
    })


from django import forms  # ✅ Required for forms.CheckboxInput
from django.shortcuts import render, get_object_or_404, redirect
from admission.models import PUAdmission
from admission.forms import PUAdmissionForm
from core.utils import get_logged_in_user, log_activity

@custom_login_required
def view_pu_admission(request, pk):
    admission = get_object_or_404(PUAdmission, pk=pk)
    form = PUAdmissionForm(instance=admission)

    # ✅ Disable all form fields
    for name, field in form.fields.items():
        if isinstance(field.widget, forms.CheckboxInput):
            field.widget.attrs['disabled'] = True
        else:
            field.widget.attrs.update({'readonly': True, 'disabled': True})


           
    # ✅ Prepare subject rows
    subject_rows = []
    for i in range(1, 7):
        percent_field = f"total_marks_percentage{i}"
        subject_rows.append({
            "subject_field": form[f"subject{i}"],
            "max_marks_field": form[f"max_marks{i}"],
            "marks_field": form[f"marks_obtained{i}"],
            "percent_field": form[percent_field],
            "subject_val": getattr(admission, f"subject{i}", ""),
            "max_marks_val": getattr(admission, f"max_marks{i}", ""),
            "marks_val": getattr(admission, f"marks_obtained{i}", ""),
            "percent_val": getattr(admission, percent_field, ""),
        })

        user = get_logged_in_user(request)
        log_activity(user, 'viewed', admission)

    return render(request, 'admission/admission_form.html', {
        'form': form,
        'admission': admission,
        'subject_rows': subject_rows,
        'add_mode': False,
        'edit_mode': False,
        'view_mode': True,
    })


from django.utils.http import url_has_allowed_host_and_scheme

@custom_login_required
def edit_pu_admission(request, pk):
    admission = get_object_or_404(PUAdmission, pk=pk)
    success = False

    # Get the 'next' parameter from the query string (e.g., ?next=/pending-applications/)
    next_url = request.GET.get('next')

    if request.method == 'POST':
        form = PUAdmissionForm(request.POST, request.FILES, instance=admission)
        if form.is_valid():
            pu_admission = form.save()

            # ✅ Log the update
            user = get_logged_in_user(request)
            log_activity(user, 'edited', pu_admission)

            success = True

            # ✅ Redirect to 'next' if it’s safe, else fallback
            if next_url and url_has_allowed_host_and_scheme(next_url, request.get_host()):
                return redirect(next_url)
            return HttpResponseRedirect(reverse('admission_list') + '?success=edit')
        else:
            print("❌ Form errors:", form.errors)
    else:
        form = PUAdmissionForm(instance=admission)

    # ✅ Prepare subject rows
    subject_rows = []
    for i in range(1, 7):
        percent_field = f"total_marks_percentage{i}"
        subject_rows.append({
            "subject_field": form[f"subject{i}"],
            "max_marks_field": form[f"max_marks{i}"],
            "marks_field": form[f"marks_obtained{i}"],
            "percent_field": form[percent_field],
            "subject_val": getattr(admission, f"subject{i}", ""),
            "max_marks_val": getattr(admission, f"max_marks{i}", ""),
            "marks_val": getattr(admission, f"marks_obtained{i}", ""),
            "percent_val": getattr(admission, percent_field, ""),
        })

    return render(request, 'admission/admission_form.html', {
        'form': form,
        'success': success,
        'edit_mode': True,
        'add_mode': False,
        'view_mode': False,
        'admission': admission,
        'next_admission_no': admission.admission_no,
        'subject_rows': subject_rows,
    })


from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import PUAdmission

@custom_login_required
def delete_pu_admission(request, pk):
    admission = get_object_or_404(PUAdmission, pk=pk)

    # Store before delete
    admission_no = admission.admission_no
    student_name = admission.student_name

    admission.delete()

    # ✅ Log the delete
    user = get_logged_in_user(request)
    log_activity(user, 'deleted', admission)

    # ✅ Success message with info
    messages.success(request, f"Admission {admission_no} deleted successfully for {student_name}.")
    return redirect('admission_list')  # Update if your URL name differs
# ---------- Degree Admission Form View ----------
import datetime
import requests
from django.shortcuts import render
from .models import DegreeAdmission, Enquiry1
from .forms import DegreeAdmissionForm



from django.shortcuts import render
from .models import DegreeAdmission
from master.models import UserCustom

@custom_login_required
def degree_admission_list(request):
    admissions = DegreeAdmission.objects.all().select_related('course')

    # Sort by admission serial number
    def extract_serial(adm):
        try:
            return int(adm.admission_no.split('/')[-1])
        except:
            return 0

    admissions = sorted(admissions, key=extract_serial)

    # Map user IDs to usernames
    user_ids = [a.admission_taken_by for a in admissions if a.admission_taken_by]
    users = UserCustom.objects.filter(id__in=user_ids).values('id', 'username')
    user_dict = {user['id']: user['username'] for user in users}

    for admission in admissions:
        admission.created_by_username = user_dict.get(admission.admission_taken_by, "Unknown")

        # Label based on admission_source saved in DB
        if admission.admission_source == 'enquiry':
            admission.taken_from_label = "From Enquiry"
        elif admission.admission_source == 'bulk_import':
            admission.taken_from_label = "From Bulk"
        else:
            admission.taken_from_label = "Direct"

        # You do not need to attach admission_source separately; 
        # use admission.get_admission_source_display in template.

    return render(request, 'admission/degree_admission_list.html', {'admissions': admissions})


# ---------- Degree Admission Form View ----------

import re
from datetime import date
from django.shortcuts import render
from django.contrib import messages
from .models import DegreeAdmission, Enquiry2
from .forms import DegreeAdmissionForm
from .utils import generate_next_receipt_no_shared
from master.models import FeeMaster
from core.utils import get_logged_in_user, log_activity


def match_fee_name(name):
    name = name.strip().lower().replace(" ", "")
    if any(prefix in name for prefix in ["tution", "tuition"]):
        return "tuition_fee"
    elif "application" in name:
        return "application_fee"
    elif "books" in name:
        return "books_fee"
    elif "uniform" in name:
        return "uniform_fee"
    elif "transport" in name:
        return "transport_fees"
    return None

from django.db.models import Q
from django.shortcuts import render, redirect
from datetime import date

@custom_login_required
def degree_admission_form(request, enquiry_no=None):
    success = False
    prefix = "PSCM2025-26DG"
 
    last = DegreeAdmission.objects.filter(admission_no__startswith=f"{prefix}/").order_by('-id').first()
    last_num = int(re.search(r'/(\d+)$', last.admission_no).group(1)) if (last and last.admission_no) else 0
    next_serial = last_num + 1
    next_adm_no = f"{prefix}/{next_serial:03d}"
    next_rcp_no, last_receipt_no = generate_next_receipt_no_shared()
 
 
    marks_initial = {f"{key}{i}": "" for i in range(1, 7) for key in ["subject", "max_marks", "marks_obtained", "total_marks_percentage"]}
 
    tuition_fee = books_fee = uniform_fee = application_fee = transport_fees = 0
    course_type_id = course_id = None
 
    # ✅ Default course_type as "School of Commerce"
    default_course_type = CourseType.objects.filter(
    Q(name__icontains="School of Commerce") |
    Q(name__icontains="Commerce") |
    Q(name__icontains="BCom")).first()
 
    if enquiry_no:
        try:
            en = Enquiry2.objects.get(enquiry_no=enquiry_no)
            course_type_id = en.course_type.id if en.course_type else None
            course_id = en.course.id if hasattr(en, 'course') else None
        except Enquiry2.DoesNotExist:
            pass
    elif default_course_type:
        course_type_id = default_course_type.id
 
    if request.method == "POST":
        form = DegreeAdmissionForm(request.POST, request.FILES)
        if form.is_valid():
            ct = form.cleaned_data.get("course_type")
            c = form.cleaned_data.get("course")
            course_type_id = ct.id if ct else None
            course_id = c.id if c else None
 
            if course_type_id and course_id:
                qs = FeeMaster.objects.filter(program_type_id=course_type_id, combination_id=course_id)
                for fee in qs:
                    key = match_fee_name(fee.fee_name)
                    if key:
                        locals()[key] = fee.fee_amount
    else:
        form = DegreeAdmissionForm(initial={
            **marks_initial,
            'admission_no': next_adm_no,
            'receipt_no': next_rcp_no,
            'course_type': course_type_id  # ✅ Set default course_type
        })
 
        if course_type_id and course_id:
            qs = FeeMaster.objects.filter(program_type_id=course_type_id, combination_id=course_id)
            for fee in qs:
                key = match_fee_name(fee.fee_name)
                if key:
                    locals()[key] = fee.fee_amount
 
    if request.method == "POST" and form.is_valid():
        da = form.save(commit=False)
        da.admission_no = next_adm_no
        da.receipt_no = next_rcp_no
        da.receipt_date = date.today()
 
        ef = form.cleaned_data.get("enquiry_no")
        sn = form.cleaned_data.get("student_name")
        if not enquiry_no and not ef:
            tmp = Enquiry2.objects.filter(student_name=sn).first()
            enquiry_no = tmp.enquiry_no if tmp else None
        da.enquiry_no = enquiry_no or ef
 
        da.admission_taken_by = request.session.get('user_id')
        da.admission_source = 'enquiry' if ef else 'direct'
        da.admission_date = date.today() if not da.admission_date else da.admission_date
 
        da.application_fee = application_fee
        da.tuition_fee = tuition_fee
        da.books_fee = books_fee
        da.uniform_fee = uniform_fee
        if hasattr(da, "transport_fees"):
            da.transport_fees = transport_fees
 
        da.final_fee_after_advance = tuition_fee - (da.tuition_advance_amount or 0) - (da.scholarship_amount or 0)
        da.save()
 
        log_activity(get_logged_in_user(request), 'created', da)
 
        messages.success(
            request,
            f"Admission {da.admission_no} successfully added for {da.student_name}. "
            f"Receipt No: {da.receipt_no}, Date: {da.receipt_date.strftime('%d-%m-%Y')}"
        )
        return HttpResponseRedirect (reverse('degree_admission_list') + '?success=add')
 
 
        success = True
        next_serial += 1
        next_adm_no = f"{prefix}/{next_serial:03d}"
        next_rcp_no, _ = generate_next_receipt_no_shared()
        form = DegreeAdmissionForm(initial={**marks_initial, 'admission_no': next_adm_no, 'receipt_no': next_rcp_no})
 
    else:
        initial = {
            **marks_initial,
            'admission_no': next_adm_no,
            'receipt_no': next_rcp_no,
            'course_type': course_type_id,  # ✅ Ensure it’s included even outside POST
            'receipt_date': date.today(),  # 👈 add this line
 
        }
        if enquiry_no:
            try:
                en = Enquiry2.objects.get(enquiry_no=enquiry_no)
                initial.update({
                    'enquiry_no': en.enquiry_no,
                    'student_name': en.student_name,
                    'gender': en.gender,
                    'parent_name': en.parent_name,
                    'parent_mobile_no': en.parent_phone,
                    'email': en.email,
                    'sslc_percentage': en.percentage_12th,
                    'course': course_id,
                })
            except Enquiry2.DoesNotExist:
                pass
        form = DegreeAdmissionForm(initial=initial)
 
    subject_rows = [{
        "subject_field": form[f"subject{i}"],
        "max_marks_field": form[f"max_marks{i}"],
        "marks_field": form[f"marks_obtained{i}"],
        "percent_field": form[f"total_marks_percentage{i}"],
        "subject_val": "", "max_marks_val": "", "marks_val": "", "percent_val": ""
    } for i in range(1, 7)]
 
    return render(request, 'admission/degree_admission_form.html', {
        'form': form,
        'form_submission_success': success,
        'next_admission_no': next_adm_no,
        'subject_rows': subject_rows,
        'add_mode': True,
        'edit_mode': False,
        'view_mode': False,
        'last_receipt_no': last_receipt_no if next_rcp_no else None,
 
    })

@custom_login_required
def ajax_load_courses(request):
    course_type_id = request.GET.get("course_type")
    courses = Course.objects.filter(course_type_id=course_type_id).values('id', 'name')
    return JsonResponse(list(courses), safe=False)

@custom_login_required
def view_degree_admission(request, pk):
    admission = get_object_or_404(DegreeAdmission, pk=pk)
    form = DegreeAdmissionForm(instance=admission)

    # Disable all fields in the form for view mode
    for field in form.fields.values():
        field.disabled = True

    # Prepare subject rows
    subject_rows = []
    for i in range(1, 7):
        # Handle missing percentage field gracefully
        percent_field_name = f"total_marks_percentage{i}"
        percent_field = form[percent_field_name] if percent_field_name in form.fields else None
        percent_val = getattr(admission, percent_field_name, "")

        subject_rows.append({
            "subject_field": form[f"subject{i}"],
            "max_marks_field": form[f"max_marks{i}"],
            "marks_field": form[f"marks_obtained{i}"],
            "percent_field": percent_field,
            "subject_val": getattr(admission, f"subject{i}", ""),
            "max_marks_val": getattr(admission, f"max_marks{i}", ""),
            "marks_val": getattr(admission, f"marks_obtained{i}", ""),
            "percent_val": percent_val,
        })

    context = {
        "form": form,
        "admission": admission,
        "subject_rows": subject_rows,
        "add_mode": False,
        "edit_mode": False,
        "view_mode": True,
    }
    return render(request, "admission/degree_admission_form.html", context)


from django.utils.http import url_has_allowed_host_and_scheme

@custom_login_required
def edit_degree_admission(request, pk):
    admission = get_object_or_404(DegreeAdmission, pk=pk)
    view_only = request.GET.get("view") == "1"
    next_url = request.GET.get("next")  # 🔁 Flexible redirect target

    if request.method == 'POST' and not view_only:
        form = DegreeAdmissionForm(request.POST, request.FILES, instance=admission)
        if form.is_valid():
            form.save()
            user = get_logged_in_user(request)
            log_activity(user, 'edited', admission)

            messages.success(request, f"Admission No '{admission.admission_no}' updated successfully.")

            # ✅ Redirect to next if valid, fallback to degree_admission_list
            if next_url and url_has_allowed_host_and_scheme(next_url, request.get_host()):
                return redirect(next_url)
            return redirect('degree_admission_list')
    else:
        form = DegreeAdmissionForm(instance=admission)

    if view_only:
        for field in form.fields.values():
            field.disabled = True

    # Build subject_rows for template rendering
    subject_rows = []
    for i in range(1, 7):
        subject_rows.append({
            "subject_field": form[f"subject{i}"],
            "max_marks_field": form[f"max_marks{i}"],
            "marks_field": form[f"marks_obtained{i}"],
            "percent_field": form[f"total_marks_percentage{i}"],
            "subject_val": getattr(admission, f"subject{i}", ""),
            "max_marks_val": getattr(admission, f"max_marks{i}", ""),
            "marks_val": getattr(admission, f"marks_obtained{i}", ""),
            "percent_val": getattr(admission, f"total_marks_percentage{i}", ""),
        })

    return render(request, 'admission/degree_admission_form.html', {
        'form': form,
        'edit_mode': not view_only,
        'view_mode': view_only,
        'admission': admission,
        'subject_rows': subject_rows,
        'add_mode': False,
    })

from django.shortcuts import redirect
from django.contrib import messages
from .models import DegreeAdmission

@custom_login_required
def delete_degree_admission(request, pk):
    admission = DegreeAdmission.objects.filter(pk=pk).first()

    if admission:
        admission_no = admission.admission_no
        user = get_logged_in_user(request)
        log_activity(user, 'deleted', admission)
        admission.delete()

        messages.success(request, f"Admission No '{admission_no}' deleted successfully.")
    else:
        messages.warning(request, "This admission record does not exist or was already deleted.")

    return redirect('degree_admission_list')


from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import PUAdmission, DegreeAdmission, PUAdmissionshortlist, DegreeAdmissionshortlist
import json

@custom_login_required
def shortlisted_students_view(request):
    stream = request.GET.get('stream', 'PU')

    if stream == 'PU':
        students = PUAdmission.objects.filter(
            Q(application_status__iexact='Shortlisted') |
            Q(application_status__iexact='Shortlisted Management') |
            Q(application_status__iexact='Shortlisted for Management')
        )
        approved_ids = list(
            PUAdmissionshortlist.objects.filter(admin_approved=True).values_list('admission_no', flat=True)
        )
        not_approved_ids = list(
            PUAdmissionshortlist.objects.filter(admin_approved=False).values_list('admission_no', flat=True)
        )

    elif stream == 'Degree':
        students = DegreeAdmission.objects.filter(
            Q(application_status__iexact='Shortlisted') |
            Q(application_status__iexact='Shortlisted Management') |
            Q(application_status__iexact='Shortlisted for Management')
        )
        approved_ids = list(
            DegreeAdmissionshortlist.objects.filter(admin_approved=True).values_list('admission_no', flat=True)
        )
        not_approved_ids = list(
            DegreeAdmissionshortlist.objects.filter(admin_approved=False).values_list('admission_no', flat=True)
        )

    else:
        students = []
        approved_ids = []
        not_approved_ids = []

    context = {
        'stream': stream,
        'students': students,
        'approved_ids': approved_ids,
        'not_approved_ids': not_approved_ids,
    }
    return render(request, 'admission/shortlisted_students.html', context)


@custom_login_required
@csrf_exempt
def approve_student(request, stream, student_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            approved = bool(int(data.get('approved', 0)))

            if stream == 'PU':
                student = get_object_or_404(PUAdmission, id=student_id)
                shortlist, created = PUAdmissionshortlist.objects.get_or_create(
                    admission_no=student.admission_no,
                    defaults={'quota_type': student.quota_type,
                              'student_name': student.student_name  # Add this
                              }

                )
                shortlist.admin_approved = approved
                shortlist.save()

            elif stream == 'Degree':
                student = get_object_or_404(DegreeAdmission, id=student_id)
                shortlist, created = DegreeAdmissionshortlist.objects.get_or_create(
                    admission_no=student.admission_no,
                    defaults={'quota_type': student.quota_type,
                              'student_name': student.student_name  # Add this
                              }
                )
                shortlist.admin_approved = approved
                shortlist.save()

            else:
                return JsonResponse({'status': 'error', 'message': 'Invalid stream'}, status=400)

            return JsonResponse({'status': 'success', 'approved': shortlist.admin_approved})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


# from django.shortcuts import render, redirect
# from .forms import enquiryform
# from .models import enquiry

# from .models import enquiry
# from .forms import enquiryform

# def enquiry_form_view(request):
#     success = false

#     if request.method == 'post':
#         # fetch the latest enquiry again at the time of post
#         last_enquiry = enquiry.objects.order_by('id').last()
#         if last_enquiry and last_enquiry.enquiry_no and last_enquiry.enquiry_no.startswith("enq-"):
#             try:
#                 last_number = int(last_enquiry.enquiry_no.split('-')[1])
#             except (indexerror, valueerror):
#                 last_number = 0
#             next_enquiry_no = f"enq-{last_number + 1:03d}"
#         else:
#             next_enquiry_no = "enq-001"

#         form = enquiryform(request.post)
#         if form.is_valid():
#             enquiry = form.save(commit=false)
#             enquiry.enquiry_no = next_enquiry_no
#             enquiry.save()
#             success = true

#             # now prepare next enquiry no for fresh form
#             next_number = int(next_enquiry_no.split('-')[1]) + 1
#             next_enquiry_no = f"enq-{next_number:03d}"
#             form = enquiryform(initial={'enquiry_no': next_enquiry_no})
#     else:
#         # for get method only
#         last_enquiry = enquiry.objects.order_by('id').last()
#         if last_enquiry and last_enquiry.enquiry_no and last_enquiry.enquiry_no.startswith("enq-"):
#             try:
#                 last_number = int(last_enquiry.enquiry_no.split('-')[1])
#             except (indexerror, valueerror):
#                 last_number = 0
#             next_enquiry_no = f"enq-{last_number + 1:03d}"
#         else:
#             next_enquiry_no = "enq-001"

#         form = enquiryform(initial={'enquiry_no': next_enquiry_no})

#     return render(request, 'admission/enquiry_form.html', {
#         'form': form,
#         'success': success
#     })
from django.shortcuts import render
from .models import PUAdmission, DegreeAdmission, CourseType
from django.http import HttpResponseRedirect

@custom_login_required
def shortlist_display(request):
    selection = request.GET.get('type', 'PU')  # Default to PU
    course_type_id = request.GET.get('course_type')  # ID from dropdown
    course_types = CourseType.objects.all().order_by('name')  # For dropdown

    shortlisted = []

    if selection == 'PU':
        queryset = PUAdmission.objects.filter(status='Confirmed')
        if course_type_id:
            queryset = queryset.filter(course_type__id=course_type_id)
        shortlisted = queryset.order_by('id')

    elif selection == 'Degree':
        queryset = DegreeAdmission.objects.filter(status='Confirmed')
        if course_type_id:
            queryset = queryset.filter(course_type__id=course_type_id)
        shortlisted = queryset.order_by('id')

    return render(request, 'admission/shortlist_display.html', {
        'shortlisted': shortlisted,
        'selection': selection,
        'course_types': course_types,
        'selected_course_type': int(course_type_id) if course_type_id else None
    })



from django.shortcuts import render, get_object_or_404, redirect
from .models import PUFeeDetail, DegreeFeeDetail, PUAdmission, DegreeAdmission
from .forms import PUFeeDetailForm, DegreeFeeDetailForm

from django.shortcuts import render, get_object_or_404, redirect
from .models import PUAdmission, PUFeeDetail
from .forms import PUFeeDetailForm

@custom_login_required
def pu_fee_detail_form(request, admission_id):
    admission = get_object_or_404(PUAdmission, pk=admission_id, status="Confirmed")
    fee = PUFeeDetail.objects.filter(admission_no=admission.admission_no).first()

    # If editing, show the instance; if creating, set initial data
    if fee:
        form = PUFeeDetailForm(instance=fee)
    else:
        initial_data = {
            'tuition_fee': admission.tuition_fee,
            'scholarship': admission.scholarship_amount,
            'transport_fee': admission.transport_amount,
            'hostel_fee': getattr(admission, 'hostel_amount', 0),
            'books_fee': admission.books_fee,
            'uniform_fee': admission.uniform_fee,
            'tuition_advance_amount': admission.tuition_advance_amount,
            'payment_mode': getattr(admission, 'payment_mode', None),
            'final_fee_after_advance': admission.final_fee_after_advance,
        }
        # Mutually exclusive hostel/transport fee
        if initial_data.get('hostel_fee'):
            initial_data['transport_fee'] = 0
        elif initial_data.get('transport_fee'):
            initial_data['hostel_fee'] = 0
        form = PUFeeDetailForm(initial=initial_data)

    if request.method == 'POST':
        form = PUFeeDetailForm(request.POST, instance=fee)
        if form.is_valid():
            fee_detail = form.save(commit=False)
            fee_detail.admission_no = admission.admission_no
            fee_detail.student_name = admission.student_name
            fee_detail.course = admission.course

            # Mutually exclusive hostel/transport fee on save
            if (fee_detail.hostel_fee or 0) > 0:
                fee_detail.transport_fee = 0
            elif (fee_detail.transport_fee or 0) > 0:
                fee_detail.hostel_fee = 0

            tuition_fee = fee_detail.tuition_fee or 0
            scholarship = fee_detail.scholarship or 0
            advance = fee_detail.tuition_advance_amount or 0
            transport = fee_detail.transport_fee or 0
            hostel = fee_detail.hostel_fee or 0
            books = fee_detail.books_fee or 0
            uniform = fee_detail.uniform_fee or 0

            fee_detail.gross_fee = tuition_fee + transport + hostel + books + uniform
            fee_detail.tuition_paid = advance + scholarship
            fee_detail.final_fee_after_advance = tuition_fee - fee_detail.tuition_paid

            fee_detail.save()
            return redirect('shortlist_display')

    return render(request, 'admission/fee_detail_form.html', {
        'form': form,
        'admission': admission,
        'type': 'PU',
        'form_title': 'PU Fee Detail Form',
    })

from django.shortcuts import render, get_object_or_404, redirect
from .models import DegreeAdmission, DegreeFeeDetail
from .forms import DegreeFeeDetailForm

@custom_login_required
def degree_fee_detail_form(request, admission_id):
    admission = get_object_or_404(DegreeAdmission, pk=admission_id, status="Confirmed")
    fee = DegreeFeeDetail.objects.filter(admission_no=admission.admission_no).first()

    # Initial logic for form display
    if fee:
        form = DegreeFeeDetailForm(instance=fee)
    else:
        initial_data = {
            'tuition_fee': admission.tuition_fee,
            'scholarship': admission.scholarship_amount,
            'transport_fee': admission.transport_amount,
            'hostel_fee': getattr(admission, 'hostel_amount', 0),
            'books_fee': admission.books_fee,
            'uniform_fee': admission.uniform_fee,
            'tuition_advance_amount': admission.tuition_advance_amount,
            'payment_mode': getattr(admission, 'payment_mode', None),
            # SHOW THE FINAL FEE (read-only) in the fee form
            'final_fee_after_advance': admission.final_fee_after_advance,
        }
        # Mutually exclusive hostel/transport fee
        if initial_data.get('hostel_fee'):
            initial_data['transport_fee'] = 0
        elif initial_data.get('transport_fee'):
            initial_data['hostel_fee'] = 0
        form = DegreeFeeDetailForm(initial=initial_data)

    if request.method == 'POST':
        form = DegreeFeeDetailForm(request.POST, instance=fee)
        if form.is_valid():
            fee_detail = form.save(commit=False)
            fee_detail.admission_no = admission.admission_no
            fee_detail.student_name = admission.student_name
            fee_detail.course = admission.course

            # Mutually exclusive hostel/transport fee on save
            if (fee_detail.hostel_fee or 0) > 0:
                fee_detail.transport_fee = 0
            elif (fee_detail.transport_fee or 0) > 0:
                fee_detail.hostel_fee = 0

            tuition_fee = fee_detail.tuition_fee or 0
            scholarship = fee_detail.scholarship or 0
            advance = fee_detail.tuition_advance_amount or 0
            transport = fee_detail.transport_fee or 0
            hostel = fee_detail.hostel_fee or 0
            books = fee_detail.books_fee or 0
            uniform = fee_detail.uniform_fee or 0

            fee_detail.gross_fee = tuition_fee + transport + hostel + books + uniform
            # Always recalculate to keep in sync (ignore incoming POST value)
            fee_detail.final_fee_after_advance = tuition_fee - scholarship - advance

            fee_detail.save()
            return redirect('shortlist_display')
    return render(request, 'admission/fee_detail_form.html', {
        'form': form,
        'admission': admission,
        'type': 'Degree',
        'form_title': 'Degree Fee Detail Form',
    })

from django.shortcuts import render, redirect

from django.http import JsonResponse

from .forms import Enquiry1Form

from .models import Enquiry1, Course

from django.contrib import messages
 
from django.db.models import Value, CharField

from django.utils import timezone
 
@custom_login_required
def enquiry_list1(request):

    # Get current date and time

    now = timezone.now()
 
    # Calculate the start of the current month

    start_of_month = now.replace(day=1)
 
    # Calculate the end of the current month

    # This is the first day of the next month

    next_month = start_of_month.replace(month=now.month % 12 + 1)

    end_of_month = next_month - timedelta(days=1)
 
    # Filter Enquiries for this month

    enquiries1 = Enquiry1.objects.filter(enquiry_date__range=[start_of_month, end_of_month]).annotate(

        enquiry_type=Value('PU', output_field=CharField())

    )

    enquiries2 = Enquiry2.objects.filter(enquiry_date__range=[start_of_month, end_of_month]).annotate(

        enquiry_type=Value('DEG', output_field=CharField())

    )
 
    # Combine both enquiry sets

    enquiries = list(enquiries1) + list(enquiries2)
 
    # Extract all user IDs where admission_taken_by is not None

    user_ids = [a.created_by for a in enquiries if a.created_by]
 
    # Fetch corresponding users with their usernames

    users = UserCustom.objects.filter(id__in=user_ids).values('id', 'username')
 
    # Create a dictionary to map user IDs to usernames

    user_dict = {user['id']: user['username'] for user in users}
 
    # Add the username dynamically to each enquiry object

    for enquiry in enquiries:

        enquiry.created_by_username = user_dict.get(enquiry.created_by, "Unknown")
 
    # Render the enquiry list template with the filtered enquiries

    return render(request, 'admission/enquiry_list1.html', {

        'enquiries': enquiries,

    })
 
 





from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from .models import Enquiry1, Enquiry2
from .utils import send_msgkart_template
from django.core.exceptions import ObjectDoesNotExist

from core.utils import get_logged_in_user, log_activity

# @csrf_exempt

# from django.http import JsonResponse
# from django.views.decorators.csrf import csrf_exempt
# from django.shortcuts import get_object_or_404
# from .models import Enquiry1, Enquiry2
# from .utils import send_msgkart_template
# from django.core.exceptions import ObjectDoesNotExist

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.utils.timezone import now
from django.core.exceptions import ObjectDoesNotExist
from admission.models import Enquiry1, Enquiry2
from .utils import send_msgkart_template  # adjust this import as needed

@custom_login_required
@csrf_exempt
def send_whatsapp_message(request, enquiry_no):

    user = get_logged_in_user(request)

    if request.method == "POST":
        # Try Enquiry1, then Enquiry2
        try:
            enquiry = Enquiry1.objects.get(enquiry_no=enquiry_no)
            enquiry_model = "PU"
        except Enquiry1.DoesNotExist:
            try:
                enquiry = Enquiry2.objects.get(enquiry_no=enquiry_no)
                enquiry_model = "DEG"
            except Enquiry2.DoesNotExist:
                return JsonResponse({"status": "error", "message": "Enquiry not found."}, status=404)

        phone = (enquiry.parent_phone or "").strip().replace(" ", "")
        if not phone.startswith("+91"):
            phone = "+91" + phone

        try:
            course_name = enquiry.course.name if enquiry.course else "N/A"
        except ObjectDoesNotExist:
            course_name = "N/A"

        try:
            course_type_name = enquiry.course_type.name if enquiry.course_type else "N/A"
        except ObjectDoesNotExist:
            course_type_name = "N/A"

        param_list = [
            enquiry.parent_name or "N/A",
            enquiry.student_name or "N/A",
            course_name,
            course_type_name,
            enquiry.enquiry_no or "N/A",
            enquiry.enquiry_date.strftime("%d-%m-%Y") if enquiry.enquiry_date else "N/A",
            enquiry.email or "N/A",
            enquiry.gender or "N/A",
            enquiry.guardian_relation or "N/A",
            enquiry.parent_phone or "N/A",
        ]

        try:
            response = send_msgkart_template(phone.replace("+", ""), param_list)
            res_json = response.json()
            transaction_id = res_json.get("transactionId", "")
            status = res_json.get("status", "")
            error_msg = res_json.get("error", "") or res_json.get("message", "")

            # Update DB status
            if response.status_code in [200, 202] and (transaction_id or (status and status.lower() == "success")):
                enquiry.whatsapp_status = 'sent'
                enquiry.whatsapp_sent_date = now().date()  # ✅ Set sent date
                enquiry.save(update_fields=['whatsapp_status', 'whatsapp_sent_date'])
                return JsonResponse({
                    "status": "success",
                    "message": "WhatsApp message submitted for delivery.",
                    "transaction_id": transaction_id or ""
                })
            else:
                enquiry.whatsapp_status = 'failed'
                enquiry.whatsapp_sent_date = None  # Optional: clear date on failure
                enquiry.save(update_fields=['whatsapp_status', 'whatsapp_sent_date'])
                return JsonResponse({
                    "status": "error",
                    "message": f"Failed to send WhatsApp message. Error: {error_msg or response.text or 'Unknown error'}",
                    "transaction_id": transaction_id or ""
                }, status=400)

        except Exception as e:
            enquiry.whatsapp_status = 'failed'
            enquiry.whatsapp_sent_date = None
            enquiry.save(update_fields=['whatsapp_status', 'whatsapp_sent_date'])

            log_activity(user, 'error', enquiry)
            return JsonResponse({
                "status": "error",
                "message": f"Error sending WhatsApp: {str(e)}"
            }, status=500)

    else:
        return JsonResponse({"status": "error", "message": "Invalid request method."}, status=405)





from django.db.models import Q
from django.utils import timezone

from django.shortcuts import render

from admission.models import Enquiry1, Enquiry2, PUAdmission, DegreeAdmission, FollowUp
 
@custom_login_required
def enquiry_dashboard(request):

    now = timezone.now()

    start_of_month = now.replace(day=1)

    start_of_week = now - timedelta(days=now.weekday())  # Monday

    end_of_week = start_of_week + timedelta(days=6)       # Sunday
 
    # Converted enquiry numbers

    pu_converted_nos = PUAdmission.objects.exclude(enquiry_no__isnull=True).exclude(enquiry_no='').values_list('enquiry_no', flat=True)

    degree_converted_nos = DegreeAdmission.objects.exclude(enquiry_no__isnull=True).exclude(enquiry_no='').values_list('enquiry_no', flat=True)
 
    # Total enquiries this month

    pu_enquiries = Enquiry1.objects.filter(enquiry_date__gte=start_of_month)

    degree_enquiries = Enquiry2.objects.filter(enquiry_date__gte=start_of_month)

    total_enquiries = pu_enquiries.count() + degree_enquiries.count()
 
    # Converted enquiries this month

    pu_converted_qs = PUAdmission.objects.filter(
        admission_date__gte=start_of_month,
        enquiry_no__isnull=False
    ).exclude(enquiry_no='').filter(
        enquiry_no__in=Enquiry1.objects.filter(is_converted=True).values_list('enquiry_no', flat=True)
    )

    degree_converted_qs = DegreeAdmission.objects.filter(
        admission_date__gte=start_of_month,
        enquiry_no__isnull=False
    ).exclude(enquiry_no='').filter(
        enquiry_no__in=Enquiry2.objects.filter(is_converted=True).values_list('enquiry_no', flat=True)
    )
    total_converted_enquiries=pu_converted_qs.count() + degree_converted_qs.count()
    
    

 
    # Follow-up Scheduled: this week, after now, and not converted

    followup_scheduled = FollowUp.objects.filter(

        follow_up_date__gte=now,


        follow_up_date__lte=end_of_week,

        status='Pending'

    ).exclude(

        Q(pu_enquiry__enquiry_no__in=pu_converted_nos) |

        Q(degree_enquiry__enquiry_no__in=degree_converted_nos)

    ).count()
 
    # Follow-up Required: enquiries this week, not converted, no follow-up at all

    followup_required = 0

    unconverted_enquiries = list(

        pu_enquiries.exclude(enquiry_no__in=pu_converted_nos).filter(enquiry_date__range=(start_of_week, end_of_week))

    ) + list(

        degree_enquiries.exclude(enquiry_no__in=degree_converted_nos).filter(enquiry_date__range=(start_of_week, end_of_week))

    )
 
    for enquiry in unconverted_enquiries:

        if isinstance(enquiry, Enquiry1):

            has_followup = FollowUp.objects.filter(pu_enquiry=enquiry).exists()

        else:

            has_followup = FollowUp.objects.filter(degree_enquiry=enquiry).exists()

        if not has_followup:

            followup_required += 1
 
    # Total Follow-ups Due = scheduled (future in week) + required (no follow-up)

    followups_due = followup_scheduled + followup_required
 
    # Pending follow-ups: this month, missed (before now), and not converted

    pending_followups = FollowUp.objects.filter(

        follow_up_date__lt=now,

        follow_up_date__gte=start_of_month,

        status='Pending'

    ).exclude(

        Q(pu_enquiry__enquiry_no__in=pu_converted_nos) |

        Q(degree_enquiry__enquiry_no__in=degree_converted_nos)

    ).count()
 
    # Conversion rate

    conversion_rate = (total_converted_enquiries / total_enquiries) * 100 if total_enquiries else 0
 
    context = {

        'followups_due': followups_due,

        'pending_followups': pending_followups,

        'total_enquiries': total_enquiries,

        'total_converted_enquiries': total_converted_enquiries,

        'conversion_rate': round(conversion_rate, 2),

    }
 
    return render(request, 'admission/enquiry_dashboard.html', context)
 

from django.db.models import Exists, OuterRef, Q


from django.shortcuts import render
from django.utils import timezone
from datetime import timedelta
from .models import PUAdmission, DegreeAdmission, Enquiry1, Enquiry2, FollowUp
from django.db.models import Q
from .forms import FollowUpForm

@custom_login_required
def followups_due_list(request):
    now = timezone.now()
    start_of_week = now - timedelta(days=now.weekday())
    end_of_week = start_of_week + timedelta(days=6)

    pu_converted_nos = PUAdmission.objects.exclude(enquiry_no__isnull=True).exclude(enquiry_no='').values_list('enquiry_no', flat=True)
    degree_converted_nos = DegreeAdmission.objects.exclude(enquiry_no__isnull=True).exclude(enquiry_no='').values_list('enquiry_no', flat=True)

    # Scheduled follow-ups
    followups_scheduled = FollowUp.objects.filter(
        follow_up_date__gte=now,
        follow_up_date__lte=end_of_week,
        status='Pending'
    ).exclude(
        Q(pu_enquiry__enquiry_no__in=pu_converted_nos) |
        Q(degree_enquiry__enquiry_no__in=degree_converted_nos)
    )

    # Convert to list and tag type
    scheduled_list = [{'type': 'scheduled', 'data': f} for f in followups_scheduled]

    # Enquiries needing follow-up
    followups_required = []

    pu_enquiries = Enquiry1.objects.filter(enquiry_date__range=(start_of_week, end_of_week)).exclude(enquiry_no__in=pu_converted_nos)
    degree_enquiries = Enquiry2.objects.filter(enquiry_date__range=(start_of_week, end_of_week)).exclude(enquiry_no__in=degree_converted_nos)

    for enquiry in list(pu_enquiries) + list(degree_enquiries):
        if isinstance(enquiry, Enquiry1):
            if not FollowUp.objects.filter(pu_enquiry=enquiry).exists():
                followups_required.append({'type': 'required', 'data': enquiry})
        else:
            if not FollowUp.objects.filter(degree_enquiry=enquiry).exists():
                followups_required.append({'type': 'required', 'data': enquiry})

    combined_followups = scheduled_list + followups_required

    context = {
        'followups': combined_followups
    }
    return render(request, 'admission/followups_due_list.html', context)


@custom_login_required
def pending_followups_list(request):
    now = timezone.now()
    start_of_month = now.replace(day=1)
    enquiry_no = request.GET.get('enquiry_no') 
    pu_converted_nos = PUAdmission.objects.exclude(enquiry_no__isnull=True).exclude(enquiry_no='').values_list('enquiry_no', flat=True)
    degree_converted_nos = DegreeAdmission.objects.exclude(enquiry_no__isnull=True).exclude(enquiry_no='').values_list('enquiry_no', flat=True)

    pending_followups = FollowUp.objects.filter(
        follow_up_date__lt=now,
        follow_up_date__gte=start_of_month,
        status='Pending'
    ).exclude(
        Q(pu_enquiry__enquiry_no__in=pu_converted_nos) |
        Q(degree_enquiry__enquiry_no__in=degree_converted_nos)
    )

    context = {
        'pending_followups': pending_followups,
        'enquiry_no': enquiry_no, 
    }
    return render(request, 'admission/pending_followups_list.html', context)


@custom_login_required
def schedule_follow_up_form_add(request):
    enquiry_no = request.GET.get('enquiry_no')
    student_name = request.GET.get('student_name')
    enquiry = None
    enquiry_type = None
    action = request.GET.get('action', 'add')  # default to 'edit' if not provided
    if enquiry_no:
        if enquiry_no.startswith("PU-ENQ-"):
            enquiry = get_object_or_404(Enquiry1, enquiry_no=enquiry_no)
            enquiry_type = "PU"
        elif enquiry_no.startswith("DEG-ENQ-"):
            enquiry = get_object_or_404(Enquiry2, enquiry_no=enquiry_no)
            enquiry_type = "DEG"

    initial_data = {}
    if enquiry:
        if enquiry_type == "PU":
            initial_data['combined_enquiry'] = f"pu_{enquiry.id}"
        elif enquiry_type == "DEG":
            initial_data['combined_enquiry'] = f"deg_{enquiry.id}"
        initial_data['student_name_display'] = student_name

    if request.method == 'POST':
        form = FollowUpForm(request.POST)
        if form.is_valid():
            followup = form.save(commit=False)
            followup.status = 'Pending'
            followup.save()

            user = get_logged_in_user(request)
            log_activity(user, 'followup_scheduled', followup)

            messages.success(request, "Follow-up scheduled successfully!")
            return redirect('enquiry_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = FollowUpForm(initial=initial_data)

    return render(request, 'admission/schedule_follow_up_form.html', {
        'form': form,
        'enquiry': enquiry,
        'enquiry_no': enquiry_no,
        'action': action, 
    })

from django.http import JsonResponse
from .models import Enquiry1, Enquiry2  # adjust to your model names
 
from django.http import JsonResponse
from .models import Enquiry1, Enquiry2

@custom_login_required 
def get_student_name(request):
    value = request.GET.get('value')
    if value:
        prefix, obj_id = value.split('_', 1)
        if prefix == 'pu':
            try:
                student_name = Enquiry1.objects.get(id=obj_id).student_name
                return JsonResponse({'student_name': student_name})
            except Enquiry1.DoesNotExist:
                pass
        elif prefix == 'deg':
            try:
                student_name = Enquiry2.objects.get(id=obj_id).student_name
                return JsonResponse({'student_name': student_name})
            except Enquiry2.DoesNotExist:
                pass
    return JsonResponse({'student_name': ''})


from django.shortcuts import get_object_or_404, redirect, render
from .models import FollowUp
from .forms import FollowUpForm

@custom_login_required
def schedule_follow_up_form_edit(request, pk):
    followup = get_object_or_404(FollowUp, pk=pk)
    action = request.GET.get('action', 'edit')  # default to 'edit' if not provided
    if request.method == 'POST':
        form = FollowUpForm(request.POST, instance=followup)
        form.instance.combined_enquiry = followup.combined_enquiry
        if form.is_valid():
            form.save()
 
            user = get_logged_in_user(request)
            log_activity(user, 'followup_scheduled', followup)
            messages.success(request, "Follow-up updated successfully!")
 
            return redirect('follow_up_list')
    else:
        form = FollowUpForm(instance=followup)
         
    return render(request, 'admission/schedule_follow_up_form.html', {'form': form, 'edit': True, 'action': action})


from django.shortcuts import render, get_object_or_404
from .models import FollowUp
from .forms import FollowUpForm

@custom_login_required
def schedule_follow_up_form_view(request, pk):
    followup = get_object_or_404(FollowUp, pk=pk)
    action = request.GET.get('action', 'view')  # default to 'edit' if not provided

    form = FollowUpForm(instance=followup)
    for field in form.fields.values():
        field.disabled = True  # make all fields read-only
 
    return render(request, 'admission/schedule_follow_up_form.html', {
        'form': form,
        'view': True, 'action': action
    })
@custom_login_required
def schedule_follow_up_form_delete(request, pk):
    followup = get_object_or_404(FollowUp, pk=pk)

    followup.delete()
    user = get_logged_in_user(request)
    log_activity(user, 'followup_scheduled', followup)
    messages.success(request, "Follow-up deleted successfully!")
    return redirect('follow_up_list')




@custom_login_required
def follow_up_list(request):
    followups = FollowUp.objects.select_related('pu_enquiry', 'degree_enquiry').all().order_by('-follow_up_date')
    return render(request, 'admission/follow_up_list.html', {'followups': followups})

@custom_login_required
def update_followup_status(request, id):
    if request.method == 'POST':
        followup = get_object_or_404(FollowUp, id=id)
        new_status = request.POST.get('status')
        if new_status in ['Pending', 'Completed']:
            followup.status = new_status
            followup.save()
    return redirect('follow_up_list')  

from admission.models import Enquiry1, Enquiry2
from admission.forms import Enquiry1Form, Enquiry2Form
from django.db.models import Value, CharField
# Dummy message form for demonstration
from django.db.models import Value, CharField
from django.utils import timezone
from django.shortcuts import render, get_object_or_404
from admission.models import Enquiry1, Enquiry2, FollowUp,PUAdmission,DegreeAdmission
 
from admission.forms import Enquiry1Form, Enquiry2Form
 
@custom_login_required
def enquiry_list(request):
    enquiries1 = Enquiry1.objects.all().annotate(enquiry_type=Value('PU', output_field=CharField()))
    enquiries2 = Enquiry2.objects.all().annotate(enquiry_type=Value('DEG', output_field=CharField()))
    enquiries = list(enquiries1) + list(enquiries2)
    now = timezone.now()
 
    # Extract all user IDs where admission_taken_by is not None
   
    # Add the username dynamically to each admission object
    for enquiry in enquiries:
        enquiry.created_by_username = enquiry.created_by or "Unknown"
 
    # Attach follow-up status and conversion flag
    for enquiry in enquiries:
        # Follow-up status
        if enquiry.enquiry_type == 'PU':
            followups = FollowUp.objects.filter(pu_enquiry=enquiry, status='Pending').order_by('-follow_up_date')
        elif enquiry.enquiry_type == 'DEG':
            followups = FollowUp.objects.filter(degree_enquiry=enquiry, status='Pending').order_by('-follow_up_date')
 
        if followups.exists():
            latest = followups.first()
            enquiry.latest_followup_date = latest.follow_up_date  # ✅ Add this
            if latest.follow_up_date < now:
                enquiry.followup_status = 'Pending Follow-up'
            else:
                enquiry.followup_status = 'Follow-up Scheduled'
        else:
            enquiry.followup_status = 'Follow-up Required'
            enquiry.latest_followup_date = None  # ✅ Add fallback
 
 
        # Converted check
              # Converted check and DB update
        if enquiry.enquiry_type == 'PU':
            is_conv = PUAdmission.objects.filter(enquiry_no=enquiry.enquiry_no).exists()
            if is_conv and not enquiry.is_converted:
                enquiry.is_converted = True
                enquiry.save(update_fields=['is_converted'])
            else:
                enquiry.is_converted = is_conv
 
        elif enquiry.enquiry_type == 'DEG':
            is_conv = DegreeAdmission.objects.filter(enquiry_no=enquiry.enquiry_no).exists()
            if is_conv and not enquiry.is_converted:
                enquiry.is_converted = True
                enquiry.save(update_fields=['is_converted'])
            else:
                enquiry.is_converted = is_conv
 
 
    enquiry_no = request.GET.get('enquiry_no')
    action = request.GET.get('action')  # 'view' or None
    selected_form = None
    enquiry = None
 
    if enquiry_no and action == 'view':
        if enquiry_no.startswith("PU-ENQ-"):
            enquiry = get_object_or_404(Enquiry1, enquiry_no=enquiry_no)
            selected_form = Enquiry1Form(instance=enquiry)
        elif enquiry_no.startswith("DEG-ENQ-"):
            enquiry = get_object_or_404(Enquiry2, enquiry_no=enquiry_no)
            selected_form = Enquiry2Form(instance=enquiry)
 
        if selected_form:
            for field in selected_form.fields.values():
                field.widget.attrs['disabled'] = True
 
    return render(request, 'admission/enquiry_list.html', {
        'enquiries': enquiries,
        'selected_form': selected_form,
        'enquiry_no': enquiry_no,
        'action': action
    })
 

from master.models import UserCustom
from admission.models import Enquiry1
from admission.forms import Enquiry1Form
from django.contrib import messages
from django.shortcuts import render, redirect


from django.shortcuts import render, get_object_or_404
from admission.models import Enquiry1, Enquiry2
from admission.forms import Enquiry1Form, Enquiry2Form

from django.core.exceptions import ObjectDoesNotExist

@custom_login_required
def enquiry_form_view(request, enquiry_no):
    if enquiry_no.startswith("PU"):
        enquiry = get_object_or_404(Enquiry1, enquiry_no=enquiry_no)
        form = Enquiry1Form(instance=enquiry)
        template = 'admission/enquiry1_form.html'
    else:
        enquiry = get_object_or_404(Enquiry2, enquiry_no=enquiry_no)
        form = Enquiry2Form(instance=enquiry)
        template = 'admission/enquiry2_form.html'

    # ✅ Safe course_type queryset assignment
    form.fields['course_type'].queryset = CourseType.objects.all()

    try:
        if enquiry.course_type_id:
            form.fields['course'].queryset = Course.objects.filter(course_type_id=enquiry.course_type_id).order_by('name')
        else:
            form.fields['course'].queryset = Course.objects.none()
    except ObjectDoesNotExist:
        form.fields['course'].queryset = Course.objects.none()

    # ✅ Disable all fields after setting the querysets
    for field in form.fields.values():
        field.widget.attrs['disabled'] = True

    return render(request, template, {
        'form': form,
        'next_enquiry_no': enquiry.enquiry_no,
        'view_mode': True
    })


@custom_login_required
def enquiry_form_edit(request, enquiry_no):
    if enquiry_no.startswith("PU"):
        enquiry = get_object_or_404(Enquiry1, enquiry_no=enquiry_no)
        form_class = Enquiry1Form
        template = 'admission/enquiry1_form.html'
    else:
        enquiry = get_object_or_404(Enquiry2, enquiry_no=enquiry_no)
        form_class = Enquiry2Form
        template = 'admission/enquiry2_form.html'

    if request.method == "POST":
        form = form_class(request.POST, instance=enquiry)
        if form.is_valid():
            form.save()

            user = get_logged_in_user(request)
            log_activity(user, 'edited', enquiry)

            # Add success snackbar message including enquiry_no
            messages.success(request, f"Enquiry {enquiry_no} updated successfully!")

            return redirect('enquiry_list')  # Replace with your actual listing view name
    else:
        form = form_class(instance=enquiry)
        form.fields['enquiry_date'].widget.attrs['readonly'] = True


    return render(request, template, {
        'form': form,
        'next_enquiry_no': enquiry.enquiry_no,
        'edit_mode': True
    })

@custom_login_required
def enquiry_form_delete(request, enquiry_no):
    if enquiry_no.startswith("PU"):
        model = Enquiry1
    else:
        model = Enquiry2

    enquiry = get_object_or_404(model, enquiry_no=enquiry_no)

    # Perform the delete operation
    enquiry.delete()

    # Log the activity
    user = get_logged_in_user(request)
    log_activity(user, 'deleted', enquiry)

    # Add success snackbar message including enquiry_no
    messages.success(request, f"Enquiry {enquiry_no} deleted successfully!")

    return redirect('enquiry_list')



from master.models import UserCustom
from admission.models import Enquiry1
from admission.forms import Enquiry1Form
from django.contrib import messages
from django.shortcuts import render, redirect
from django.utils import timezone
from core.utils import get_logged_in_user, log_activity  # make sure these are imported

@custom_login_required
def enquiry_form_add(request):
    next_enquiry_no = None  # Available for both GET and POST

    # Generate enquiry number
    last_enquiry = Enquiry1.objects.order_by('-id').first()
    if last_enquiry and last_enquiry.enquiry_no and last_enquiry.enquiry_no.startswith('PU-ENQ-'):
        try:
            last_number = int(last_enquiry.enquiry_no.split('-')[2])
        except (IndexError, ValueError):
            last_number = 0
    else:
        last_number = 0
    next_enquiry_no = f"PU-ENQ-{last_number+1:02d}"

    if request.method == 'POST':
        form = Enquiry1Form(request.POST)

        if form.is_valid():
            enquiry = form.save(commit=False)
            enquiry.enquiry_no = next_enquiry_no
            enquiry.enquiry_date = timezone.now().date()

            user = get_logged_in_user(request)
            if user:
                enquiry.created_by = user.username # Or user.get_full_name()

            enquiry.save()

            # Log activity
            user = get_logged_in_user(request)
            log_activity(user, 'created', enquiry)

            print("Enquiry saved with created_by:", enquiry.created_by)

            # Add success snackbar message including enquiry_no
            messages.success(request, f"Enquiry {next_enquiry_no} saved successfully!")

            return redirect('enquiry_list')
        else:
            print("Form Errors:", form.errors)
            return render(request, 'admission/enquiry1_form.html', {
                'form': form,
                'next_enquiry_no': next_enquiry_no
            })

    else:
        form = Enquiry1Form(initial={
            'enquiry_no': next_enquiry_no,
            'enquiry_date': timezone.now().date()
        })

    return render(request, 'admission/enquiry1_form.html', {
        'form': form,
        'next_enquiry_no': next_enquiry_no
    })


from master.models import UserCustom, CourseType
from admission.models import Enquiry2
from admission.forms import Enquiry2Form
from django.contrib import messages
from django.shortcuts import render, redirect
from django.utils import timezone
from core.utils import get_logged_in_user, log_activity

@custom_login_required
def degree_enquiry_add(request):
    # Generate next enquiry number
    last_enquiry = Enquiry2.objects.order_by('-id').first()
    if last_enquiry and last_enquiry.enquiry_no and last_enquiry.enquiry_no.startswith('DEG-ENQ-'):
        try:
            last_number = int(last_enquiry.enquiry_no.split('-')[2])
        except (IndexError, ValueError):
            last_number = 0
    else:
        last_number = 0

    next_enquiry_no = f"DEG-ENQ-{last_number + 1:02d}"

    if request.method == 'POST':
        form = Enquiry2Form(request.POST)
        if form.is_valid():
            enquiry = form.save(commit=False)

            # Attach creator info
            user = get_logged_in_user(request)
            if user:
               enquiry.created_by = user.username # Or user.get_full_name()

            # Set system-generated fields
            enquiry.enquiry_no = next_enquiry_no
            enquiry.enquiry_date = timezone.now().date()
            enquiry.save()

            user = get_logged_in_user(request)
            log_activity(user, 'created', enquiry)

            # ✅ Snackbar-style message with enquiry number
            messages.success(request, f"Enquiry {next_enquiry_no} saved successfully!")
            return redirect('enquiry_list')
        else:
            print("Form Errors:", form.errors)
    else:
        form = Enquiry2Form(initial={
            'enquiry_no': next_enquiry_no,
            'enquiry_date': timezone.now().date(),
        })

    return render(request, 'admission/enquiry2_form.html', {
        'form': form,
        'next_enquiry_no': next_enquiry_no if request.method != 'POST' else None
    })


@custom_login_required
def load_courses(request):
    course_type_id = request.GET.get('course_type')
    courses = Course.objects.filter(course_type_id=course_type_id).order_by('name')
    return JsonResponse(list(courses.values('id', 'name')), safe=False)

from django.http import JsonResponse

@custom_login_required
def load_courses_degree(request):
    course_type_id = request.GET.get('course_type')

    try:
        course_type_id = int(course_type_id)
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Invalid course_type ID'}, status=400)

    courses = Course.objects.filter(course_type_id=course_type_id).order_by('name')
    return JsonResponse(list(courses.values('id', 'name')), safe=False)


@custom_login_required
def load_courses(request):
    course_type_id = request.GET.get('course_type')
    courses = Course.objects.filter(course_type_id=course_type_id).order_by('name')
    return JsonResponse(list(courses.values('id', 'name')), safe=False)


from django.conf import settings
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.models import User
from .models import PUAdmissionshortlist, DegreeAdmissionshortlist
from .email_sender import EmailSender

@custom_login_required
def send_bulk_emails(request):
    if request.method == "POST":
        provider_name = settings.EMAIL_PROVIDER_NAME
        config = settings.EMAIL_PROVIDERS.get(provider_name)
        sender = EmailSender(provider_name, config)

        pu_students = PUAdmissionshortlist.objects.filter(admin_approved=True)
        degree_students = DegreeAdmissionshortlist.objects.filter(admin_approved=True)

        success_count = 0
        for student in list(pu_students) + list(degree_students):
            email = student.email
            if not email:
                continue

            student_name = student.student_name
            username = email.split('@')[0]
            password = 'Temp@1234'

            # Create or update user with default password
            user, created = User.objects.get_or_create(username=username, email=email)
            if created:
                user.set_password(password)
                user.save()

            # ✅ Build the full login URL dynamically
            login_url = request.build_absolute_uri("http://192.168.1.143:8000//admission/student-login/")

            subject = 'Login Credentials for Student Portal'
            html_content = f"""
                <p>Dear {student_name},</p>
                <p>Welcome! Your student account has been created.</p>
                <p><strong>Login URL:</strong> <a href="{login_url}">Login Here</a></p>
                <p><strong>Username:</strong> {username}</p>
                <p><strong>Password:</strong> {password}</p>
                <p>Please change your password after login.</p>
                <p>Regards,<br>Admin</p>
            """

            if sender.send_email(email, subject, html_content):
                success_count += 1

        return HttpResponse(f"✅ {success_count} emails sent successfully.")

    return render(request, 'admission/send_bulk.html')

# yourapp/views.py

# views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from admission.models import PUAdmissionshortlist, StudentLogin

DEFAULT_PASSWORD = 'Temp@1234'  # Your fixed login password

from django.shortcuts import render, redirect
from .models import StudentLogin
from admission.models import PUAdmissionshortlist, DegreeAdmissionshortlist
from django.contrib import messages

DEFAULT_PASSWORD = "Temp@1234"  # Make sure this constant is defined

@custom_login_required
def student_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        shortlist_student = None

        # Step 1: Try fetching from PUAdmissionshortlist
        try:
            shortlist_student = PUAdmissionshortlist.objects.get(student_name=username)
        except PUAdmissionshortlist.DoesNotExist:
            pass

        # Step 2: If not found, try DegreeAdmissionshortlist
        if not shortlist_student:
            try:
                shortlist_student = DegreeAdmissionshortlist.objects.get(student_name=username)
            except DegreeAdmissionshortlist.DoesNotExist:
                messages.error(request, "Invalid student name")
                return render(request, 'admission/student_login.html')

        # Step 3: Check password
        if password != DEFAULT_PASSWORD:
            messages.error(request, "Invalid password")
            return render(request, 'admission/student_login.html')

        # Step 4: Create or get StudentLogin
        student_login, created = StudentLogin.objects.get_or_create(
            admission_no=shortlist_student.admission_no,
            defaults={
                'student_name': shortlist_student.student_name,
                'email': shortlist_student.email,
                # 'phone_number': shortlist_student.parent_mobile_no,
                # 'course': shortlist_student.quota_type,
                'password': DEFAULT_PASSWORD,
            }
        )

        # Step 5: Set session and redirect
        request.session['student_id'] = student_login.id
        return redirect('reset_password')

    return render(request, 'admission/student_login.html')


# @login_required
@custom_login_required
def reset_password(request):
    student_id = request.session.get('student_id')
    if not student_id:
        return redirect('student_login')

    student = StudentLogin.objects.get(id=student_id)

    if request.method == 'POST':
        new_password = request.POST['new_password']
        student.password = new_password
        student.is_default_password = False
        student.save()
        messages.success(request, "Password changed successfully.")
        return redirect('student_login')

    return render(request, 'admission/reset_password.html', {'student': student})


# fee py
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from .models import Student
from .forms import StudentForm
from .models import PUFeeDetail, DegreeFeeDetail
from django.template.loader import get_template
from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q


@custom_login_required
def student_list(request):
    search_query = request.GET.get('search', '').strip()
    student_list = Student.objects.all().order_by('admission_no')

    if search_query:
        student_list = student_list.filter(
            Q(admission_no__icontains=search_query) |
            Q(name__icontains=search_query)
        )

    paginator = Paginator(student_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'students': page_obj,
        'search_query': search_query,
        'page_obj': page_obj,
    }
    return render(request, 'admission/student_list.html', context)

 

# views.py
from django.http import JsonResponse
from .models import Student, PUAdmission  # Assuming PUAdmission covers both PU and degree admissions

@custom_login_required
def get_student_details(request):
    admission_no = request.GET.get('admission_no')
    data = {}

    # 1️⃣ Try Student model (if still relevant)
    student = Student.objects.filter(admission_no=admission_no).first()
    if student:
        data = {
            'student_name': student.name,
            'course': student.course,
            'tuition_fee': str(student.tuition_fee or 0),
            'scholarship': str(student.scholarship or 0),
            'tuition_advance_amount': str(student.tuition_advance_amount or 0),
            'final_fee_after_advance': str(student.final_fee_after_advance or 0),
            'transport_fee': str(student.transport_fee or 0),
            'hostel_fee': str(student.hostel_fee or 0),
            'books_fee': str(student.books_fee or 0),
            'uniform_fee': str(student.uniform_fee or 0),
        }
    else:
        # 2️⃣ NEW: Fetch from PUAdmission table (where all fee info now lives)
        pu = PUAdmission.objects.filter(admission_no=admission_no).first()
        if pu:
            # Course is a FK in PUAdmission; get name/str
            # Fee fields may be None, so use 0 fallback
            data = {
                'student_name': pu.student_name,
                'course': str(pu.course) if pu.course else "",
                'tuition_fee': str(pu.tuition_fee or 0),
                'scholarship': str(pu.scholarship_amount or 0),
                'tuition_advance_amount': str(pu.tuition_advance_amount or 0),
                'final_fee_after_advance': str(getattr(pu, 'final_fee_after_advance', 0) or 0),  # If present
                'transport_fee': str(pu.transport_amount or 0),
                'hostel_fee': str(pu.hostel_amount or 0),
                'books_fee': str(pu.books_fee or 0),
                'uniform_fee': str(pu.uniform_fee or 0),
            }
        else:
            data = {'error': 'Not found'}

    return JsonResponse(data)




from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum
from datetime import datetime
from .models import Student, StudentPaymentHistory

from django.contrib import messages

import re
from datetime import date
from django.db.models import Max
from .models import Student

@custom_login_required
def generate_new_receipt_no_and_date():
    max_receipt = Student.objects.aggregate(Max('receipt_no'))['receipt_no__max']
    if max_receipt:
        match = re.search(r'PSCN-(\d+)', max_receipt)
        if match:
            number = int(match.group(1)) + 1
        else:
            number = 1
    else:
        number = 1
    return f"PSCN-{number:03d}", date.today()


from django.contrib import messages
from django.urls import reverse

@custom_login_required
def student_fee_form_add(request):
    if request.method == "POST":
        admission_no = request.POST.get('admission_no')

        # Check if student exists
        if Student.objects.filter(admission_no=admission_no).exists():
            edit_url = reverse('student_edit', args=[admission_no])
            messages.warning(request, f"Student with Admission Number <b>{admission_no}</b> already exists. <a href='{edit_url}'>Click here to proceed payment.</a>")

            # 👇 Stay on same form page, and return with message
            next_receipt_no, today_date = generate_new_receipt_no_and_date()
            return render(request, 'admission/student_form.html', {
                'student': None,
                'history_data': {},
                'next_receipt_no': next_receipt_no,
                'today_date': today_date.strftime('%Y-%m-%d'),
            })

        # Else, proceed to create
        student = Student(admission_no=admission_no)

        receipt_no, receipt_date = generate_new_receipt_no_and_date()
        student.receipt_no = receipt_no
        student.receipt_date = receipt_date

        save_student_data_from_request(student, request)
        save_payment_history(student, request)


        user = get_logged_in_user(request)
        log_activity(user, 'created', student)

        return redirect('student_list')

    # GET
    next_receipt_no, today_date = generate_new_receipt_no_and_date()
    return render(request, 'admission/student_form.html', {
        'student': None,
        'history_data': {},
        'next_receipt_no': next_receipt_no,
        'today_date': today_date.strftime('%Y-%m-%d'),
    })



from django.shortcuts import render, get_object_or_404, redirect
from .models import Student, StudentPaymentHistory
from .forms import StudentForm
from decimal import Decimal, InvalidOperation

@custom_login_required
def safe_decimal(val):
    try:
        return Decimal(val or 0)
    except (InvalidOperation, TypeError):
        return Decimal(0)

@custom_login_required
def student_fee_form_edit(request, admission_no):
    student = get_object_or_404(Student, admission_no=admission_no)

    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            student = form.save(commit=False)

            # ✅ Generate new receipt number and date
            receipt_no, receipt_date = generate_new_receipt_no_and_date()
            student.receipt_no = receipt_no
            student.receipt_date = receipt_date

            # Get newly paid amounts
            new_tuition = safe_decimal(request.POST.get('tuition_amount'))
            new_transport = safe_decimal(request.POST.get('transport_amount'))
            new_hostel = safe_decimal(request.POST.get('hostel_amount'))
            new_books = safe_decimal(request.POST.get('books_amount'))
            new_uniform = safe_decimal(request.POST.get('uniform_amount'))
            new_other = safe_decimal(request.POST.get('other_amount'))

            # Use frontend-calculated values to avoid double addition
            student.tuition_fee_paid = safe_decimal(request.POST.get('tuition_fee_paid'))
            student.transport_fee_paid = safe_decimal(request.POST.get('transport_fee_paid'))
            student.hostel_fee_paid = safe_decimal(request.POST.get('hostel_fee_paid'))
            student.books_fee_paid = safe_decimal(request.POST.get('books_fee_paid'))
            student.uniform_fee_paid = safe_decimal(request.POST.get('uniform_fee_paid'))
            student.other_amount = safe_decimal(request.POST.get('other_amount'))

            student.tuition_pending_fee = safe_decimal(request.POST.get('tuition_pending_fee'))
            student.transport_pending_fee = safe_decimal(request.POST.get('transport_pending_fee'))
            student.hostel_pending_fee = safe_decimal(request.POST.get('hostel_pending_fee'))
            student.books_pending_fee = safe_decimal(request.POST.get('books_pending_fee'))
            student.uniform_pending_fee = safe_decimal(request.POST.get('uniform_pending_fee'))

            # Due fee dates
            student.tuition_due_date = parse_date('tuition_due_date')
            student.transport_due_date = parse_date('transport_due_date')
            student.hostel_due_date = parse_date('hostel_due_date')
            student.books_due_date = parse_date('books_due_date')
            student.uniform_due_date = parse_date('uniform_due_date')
            # student.other_due_date = parse_date('other_due_date')

            # Final fee after advance
            student.final_fee_after_advance = (
                safe_decimal(student.tuition_fee) -
                safe_decimal(student.scholarship) -
                safe_decimal(student.tuition_advance_amount)
            )

            student.save()

            user = get_logged_in_user(request)
            log_activity(user, 'updated', student)

            # Save payment history
            StudentPaymentHistory.objects.create(
                admission_no=student.admission_no,
                name=student.name,
                course=student.course,
                # fee_type="Manual Entry",

                tuition_fee=student.tuition_fee,
                tuition_fee_paid=student.tuition_fee_paid,
                tuition_pending_fee=student.tuition_pending_fee,
                tuition_amount=new_tuition,

                transport_fee=student.transport_fee,
                transport_fee_paid=student.transport_fee_paid,
                transport_pending_fee=student.transport_pending_fee,
                transport_amount=new_transport,

                hostel_fee=student.hostel_fee,
                hostel_fee_paid=student.hostel_fee_paid,
                hostel_pending_fee=student.hostel_pending_fee,
                hostel_amount=new_hostel,

                books_fee=student.books_fee,
                books_fee_paid=student.books_fee_paid,
                books_pending_fee=student.books_pending_fee,
                books_amount=new_books,

                uniform_fee=student.uniform_fee,
                uniform_fee_paid=student.uniform_fee_paid,
                uniform_pending_fee=student.uniform_pending_fee,
                uniform_amount=new_uniform,

                other_fee=student.other_fee,
                other_amount=new_other,

                tuition_due_date=student.tuition_due_date,
                transport_due_date=student.transport_due_date,
                hostel_due_date=student.hostel_due_date,
                books_due_date=student.books_due_date,
                uniform_due_date=student.uniform_due_date,
                # other_due_date=student.other_due_date,

                scholarship=student.scholarship,
                tuition_advance_amount=student.tuition_advance_amount,
                final_fee_after_advance=student.final_fee_after_advance,

                # next_installment=student.next_installment,
                # next_due_date=student.next_due_date,

                payment_method=student.payment_method,
                payment_date=student.payment_date,

                # ✅ Add new receipt info to history too (optional but useful)
                receipt_no=student.receipt_no,
                receipt_date=student.receipt_date,
                branch_code=student.branch_code
            )

            return redirect('student_list')
        else:
            print(form.errors)
    else:
        form = StudentForm(instance=student)

    return render(request, 'admission/student_form.html', {
        'form': form,
        'student': student,
        'next_receipt_no': generate_new_receipt_no_and_date()[0],
        'today_date': date.today().strftime('%Y-%m-%d'),
    })


from datetime import datetime

@custom_login_required
def save_student_data_from_request(student, request):
    def get_amount(field):
        try:
            return float(request.POST.get(field) or 0)
        except:
            return 0

    def get_date(field):
        val = request.POST.get(field)
        return datetime.strptime(val, '%Y-%m-%d').date() if val else None

    student.name = request.POST.get('name')
    student.course = request.POST.get('course')

    # Tuition
    student.tuition_fee = get_amount('tuition_fee')
    student.scholarship = get_amount('scholarship')
    student.tuition_advance_amount = get_amount('tuition_advance_amount')
    student.tuition_fee_paid = get_amount('tuition_fee_paid')
    student.tuition_amount = get_amount('tuition_amount')
    student.final_fee_after_advance = student.tuition_fee - student.scholarship - student.tuition_advance_amount
    student.tuition_pending_fee = student.final_fee_after_advance - student.tuition_fee_paid
    student.tuition_due_date = get_date('tuition_due_date')

    # Transport
    student.transport_fee = get_amount('transport_fee')
    student.transport_fee_paid = get_amount('transport_fee_paid')
    student.transport_amount = get_amount('transport_amount')
    student.transport_pending_fee = student.transport_fee - student.transport_fee_paid
    student.transport_due_date = get_date('transport_due_date')

    # Hostel
    student.hostel_fee = get_amount('hostel_fee')
    student.hostel_fee_paid = get_amount('hostel_fee_paid')
    student.hostel_amount = get_amount('hostel_amount')
    student.hostel_pending_fee = student.hostel_fee - student.hostel_fee_paid
    student.hostel_due_date = get_date('hostel_due_date')

    # Books
    student.books_fee = get_amount('books_fee')
    student.books_fee_paid = get_amount('books_fee_paid')
    student.books_amount = get_amount('books_amount')
    student.books_pending_fee = student.books_fee - student.books_fee_paid
    student.books_due_date = get_date('books_due_date')

    # Uniform
    student.uniform_fee = get_amount('uniform_fee')
    student.uniform_fee_paid = get_amount('uniform_fee_paid')
    student.uniform_amount = get_amount('uniform_amount')
    student.uniform_pending_fee = student.uniform_fee - student.uniform_fee_paid
    student.uniform_due_date = get_date('uniform_due_date')

    # Other
    student.other_fee = request.POST.get('other_fee') or ''
    student.other_amount = get_amount('other_amount')
    # student.other_due_date = get_date('other_due_date')

    # Installment and Status
    # student.next_installment = get_amount('next_installment')
    # student.next_due_date = get_date('next_due_date')
    student.status = request.POST.get('status') or ''
    student.payment_method = request.POST.get('payment_method') or ''


    student.save()


@custom_login_required
def save_payment_history(student, request):
    try:
        print("Saving payment history for:", student.admission_no)
        # print("POST fee_type:", request.POST.get('fee_type', 'Unknown'))

        StudentPaymentHistory.objects.create(
            admission_no=student.admission_no,
            name=student.name,
            course=student.course,
            # fee_type=request.POST.get('fee_type', 'Unknown'),

            tuition_fee=student.tuition_fee,
            tuition_fee_paid=student.tuition_fee_paid,
            tuition_pending_fee=student.tuition_pending_fee,
            tuition_amount=student.tuition_amount,

            transport_fee=student.transport_fee,
            transport_fee_paid=student.transport_fee_paid,
            transport_pending_fee=student.transport_pending_fee,
            transport_amount=student.transport_amount,

            hostel_fee=student.hostel_fee,
            hostel_fee_paid=student.hostel_fee_paid,
            hostel_pending_fee=student.hostel_pending_fee,
            hostel_amount=student.hostel_amount,

            books_fee=student.books_fee,
            books_fee_paid=student.books_fee_paid,
            books_pending_fee=student.books_pending_fee,
            books_amount=student.books_amount,

            uniform_fee=student.uniform_fee,
            uniform_fee_paid=student.uniform_fee_paid,
            uniform_pending_fee=student.uniform_pending_fee,
            uniform_amount=student.uniform_amount,

            other_fee=student.other_fee,
            other_amount=student.other_amount,

            # ✅ DUE DATES (Make sure your model has these fields)
            tuition_due_date=student.tuition_due_date,
            transport_due_date=student.transport_due_date,
            hostel_due_date=student.hostel_due_date,
            books_due_date=student.books_due_date,
            uniform_due_date=student.uniform_due_date,
            # other_due_date=student.other_due_date,

            scholarship=student.scholarship,
            tuition_advance_amount=student.tuition_advance_amount,
            final_fee_after_advance=student.final_fee_after_advance,

            # next_installment=student.next_installment,
            # next_due_date=student.next_due_date,
            payment_method=student.payment_method,

            receipt_no=student.receipt_no,
            receipt_date=student.receipt_date,
            branch_code=student.branch_code

        )

        print("✅ Payment history saved for:", student.admission_no)

    except Exception as e:
        print("❌ Error saving payment history:", e)






# views.py
import qrcode
from io import BytesIO
from django.http import HttpResponse
from django.views.decorators.http import require_GET

@custom_login_required
@require_GET
def generate_qr_dynamic(request):
    amount = request.GET.get("amount")
    if not amount:
        return HttpResponse("Amount is required", status=400)

    upi_id = "9483508971@ybl"  # ✅ Your UPI ID
    upi_link = f"upi://pay?pa={upi_id}&pn=Pinnacle School of Commerce & Management&am={amount}&cu=INR"
    # Optional: add student info as tx note -> &tn=StudentID-123

    qr = qrcode.make(upi_link)
    buffer = BytesIO()
    qr.save(buffer)
    buffer.seek(0)

    return HttpResponse(buffer.getvalue(), content_type="image/png")







#fee auto fetch

from django.utils.dateparse import parse_date

from django.shortcuts import render
from django.db.models import Q, F, Value as V, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce
from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q
from .models import StudentPaymentHistory


from django.http import HttpResponseNotAllowed
from django.contrib import messages
 
from .models import StudentPaymentHistory

@custom_login_required 
def student_fee_form_delete(request, admission_no):
    if request.method in ['POST', 'GET']:  # TEMPORARY for testing only
        student = get_object_or_404(Student, admission_no=admission_no)
        student.delete()


        user = get_logged_in_user(request)
        log_activity(user, 'deleted', student)


        messages.success(request, f"Student {admission_no} deleted successfully.")
        return redirect('student_list')
    return HttpResponseNotAllowed(['POST'])


@custom_login_required
def student_fee_form_view(request):
    query = request.GET.get('search')
    filter_status = request.GET.get('filter_status')
    due_date = request.GET.get('due_date')

    history_list = StudentPaymentHistory.objects.all()

    # Search by admission number or name
    if query:
        history_list = history_list.filter(
            Q(admission_no__icontains=query) |
            Q(name__icontains=query)
        )

    # Filter by payment status
    if filter_status == 'paid':
        history_list = history_list.filter(final_fee_after_advance=0)
    elif filter_status == 'pending':
        history_list = history_list.exclude(final_fee_after_advance=0)

    # Filter by due date (only if a valid date is given)
    if due_date:
        parsed_due_date = parse_date(due_date)
        if parsed_due_date:
            history_list = history_list.filter(next_due_date=parsed_due_date)

    # Annotate computed fields (ensure Decimal output)
    history_list = history_list.annotate(
        base_fee=Coalesce(F('tuition_fee'), V(0), output_field=DecimalField()),
        fee_paid=ExpressionWrapper(
            Coalesce(F('tuition_fee_paid'), V(0), output_field=DecimalField()) +
            Coalesce(F('scholarship'), V(0), output_field=DecimalField()),
            output_field=DecimalField()
        ),
        final_fee=Coalesce(F('final_fee_after_advance'), V(0), output_field=DecimalField())
    ).order_by('-id')

    # Pagination
    paginator = Paginator(history_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Context to pass to template
    context = {
        'history': page_obj,
        'search_query': query or '',
        'filter_status': filter_status or '',
        'due_date': due_date or '',
        'page_obj': page_obj
    }

    return render(request, 'admission/payment_history.html', context)






from .models import Student, StudentPaymentHistory

@custom_login_required
def save_payment(request):
    if request.method == 'POST':
        admission_no = request.POST.get('admission_no')
        amount_paid = float(request.POST.get('amount') or 0)
        payment_method = request.POST.get('payment_method')

        student = Student.objects.get(admission_no=admission_no)

        student.tuition_fee_paid += amount_paid
        student.tuition_pending_fee = student.final_fee_after_advance - student.tuition_fee_paid
        student.save()

        # 🔢 Generate receipt number and date
        receipt_no, receipt_date = generate_new_receipt_no_and_date()

        StudentPaymentHistory.objects.create(
            admission_no=student.admission_no,
            name=student.name,
            course=student.course,
            tuition_fee=student.tuition_fee,
            tuition_fee_paid=student.tuition_fee_paid,
            tuition_pending_fee=student.tuition_pending_fee,
            tuition_amount=amount_paid,
            payment_method=payment_method,
            receipt_no=receipt_no,
            receipt_date=receipt_date,
            branch_code=student.branch_code
        )

        return redirect('student_list')



import openpyxl

from django.http import HttpResponse

from .models import StudentPaymentHistory  # adjust import if needed
 
@custom_login_required
def export_payments_excel(request):

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Payment History"
 
    # Header row

    ws.append([

        'Date',

        'Admission No',

        'Name',

        'Course',

        'Amount Paid',

        'Total Fee Paid',

        'Pending Fee',

        'Payment Method'

    ])
 
    # Data rows

    payments = StudentPaymentHistory.objects.all().order_by('-id')

    for p in payments:

        ws.append([

            p.created_at.strftime('%Y-%m-%d %H:%M') if hasattr(p, 'created_at') else '',

            p.admission_no,

            p.name,

            p.course,

            p.amount,

            p.base_fee,

            p.final_fee,

            p.fee_paid,

            p.pending_fee,

            p.payment_method

        ])
 
    # Prepare response

    response = HttpResponse(

        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',

    )

    response['Content-Disposition'] = 'attachment; filename=student_payments.xlsx'

    wb.save(response)

    return response
 
#export all payments getting total paid and pending fee
from django.http import HttpResponse

import openpyxl

from django.http import HttpResponse
import openpyxl
from .models import StudentPaymentHistory  # Ensure this model matches your DB

@custom_login_required
def export_payments_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment History"

    # Header Row - Includes all required fields
    ws.append([
        'Payment Date',
        'Receipt No',
        'Receipt Date',
        'Branch Code',
        'Admission No',
        'Name',
        'Course',

        # Tuition Fee
        'Tuition Fee',
        'Tuition Fee Paid',
        'Tuition Pending Fee',
        'Tuition Amount Paid',
        'Tuition Due Date',

        # Transport
        'Transport Fee',
        'Transport Fee Paid',
        'Transport Pending Fee',
        'Transport Amount Paid',
        'Transport Due Date',

        # Hostel
        'Hostel Fee',
        'Hostel Fee Paid',
        'Hostel Pending Fee',
        'Hostel Amount Paid',
        'Hostel Due Date',

        # Books
        'Books Fee',
        'Books Fee Paid',
        'Books Pending Fee',
        'Books Amount Paid',
        'Books Due Date',

        # Uniform
        'Uniform Fee',
        'Uniform Fee Paid',
        'Uniform Pending Fee',
        'Uniform Amount Paid',
        'Uniform Due Date',

        # Other
        'Other Fee',
        'Other Amount Paid',
        'Other Due Date',

        # Scholarship / Advance
        'Scholarship',
        'Tuition Advance',
        'Final Fee After Advance',

        # Totals
        'Total Fee Paid',
        'Total Fee Pending',

        # Payment Meta
        'Payment Method',
    ])

    payments = StudentPaymentHistory.objects.all().order_by('-id')

    for p in payments:
        total_fee_paid = sum([
            p.tuition_fee_paid or 0,
            p.transport_fee_paid or 0,
            p.hostel_fee_paid or 0,
            p.books_fee_paid or 0,
            p.uniform_fee_paid or 0,
        ])

        total_pending_fee = sum([
            p.tuition_pending_fee or 0,
            p.transport_pending_fee or 0,
            p.hostel_pending_fee or 0,
            p.books_pending_fee or 0,
            p.uniform_pending_fee or 0,
        ])

        ws.append([
            p.payment_date.strftime('%Y-%m-%d') if p.payment_date else '',
            p.receipt_no or '',
            p.receipt_date.strftime('%Y-%m-%d') if p.receipt_date else '',
            p.branch_code or '',
            p.admission_no,
            p.name,
            p.course,

            p.tuition_fee,
            p.tuition_fee_paid,
            p.tuition_pending_fee,
            p.tuition_amount,
            p.tuition_due_date.strftime('%Y-%m-%d') if p.tuition_due_date else '',

            p.transport_fee,
            p.transport_fee_paid,
            p.transport_pending_fee,
            p.transport_amount,
            p.transport_due_date.strftime('%Y-%m-%d') if p.transport_due_date else '',

            p.hostel_fee,
            p.hostel_fee_paid,
            p.hostel_pending_fee,
            p.hostel_amount,
            p.hostel_due_date.strftime('%Y-%m-%d') if p.hostel_due_date else '',

            p.books_fee,
            p.books_fee_paid,
            p.books_pending_fee,
            p.books_amount,
            p.books_due_date.strftime('%Y-%m-%d') if p.books_due_date else '',

            p.uniform_fee,
            p.uniform_fee_paid,
            p.uniform_pending_fee,
            p.uniform_amount,
            p.uniform_due_date.strftime('%Y-%m-%d') if p.uniform_due_date else '',

            p.other_fee or '',
            p.other_amount or 0,

            p.scholarship or 0,
            p.tuition_advance_amount or 0,
            p.final_fee_after_advance,

            total_fee_paid,
            total_pending_fee,

            p.payment_method or '',
        ])

    # Return Excel as downloadable response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=student_payment_history.xlsx'
    wb.save(response)
    return response




from django.http import JsonResponse
from .models import Enquiry1

@custom_login_required
def enquiry_lookup(request):
    enquiry_no = request.GET.get('enquiry_no')
    form_type = request.GET.get('form_type', '').lower()
    try:
        enquiry = Enquiry1.objects.get(enquiry_no=enquiry_no)
        if form_type in enquiry.course_type.name.lower():
            return JsonResponse({
                "success": True,
                "data": {
                    "student_name": enquiry.student_name,
                    "gender": enquiry.gender,
                    "parent_name": enquiry.parent_name,
                    "parent_mobile_no": enquiry.parent_phone,
                    "address": enquiry.address,
                    "email": enquiry.email,
                    "course_type": enquiry.course_type.id,  # Use .id for <select> fields
                    "course": enquiry.course.id,
                    "sslc_percentage": enquiry.percentage_10th,  # or whatever your field is
                }
            })
        else:
            return JsonResponse({
                "success": False,
                "error": f"Provided enquiry is not for {form_type.upper()} form"
            }, status=404)
    except Enquiry1.DoesNotExist:
        return JsonResponse({
            "success": False,
            "error": "Enquiry number not found"
        }, status=404)

from django.http import JsonResponse
from .models import Enquiry1  # Or your actual enquiry model

from django.http import JsonResponse
from .models import Enquiry1  # Adjust as needed

@custom_login_required
def degree_enquiry_lookup(request):
    enquiry_no = request.GET.get('enquiry_no')
    try:
        enquiry = Enquiry1.objects.get(enquiry_no=enquiry_no)
        # Check if course_type matches "Degree" (adjust as needed)
        if "degree" in enquiry.course_type.name.lower() or enquiry.course_type.id == 2:  # Adjust logic
            return JsonResponse({
                "data": {
                    "student_name": enquiry.student_name,
                    "gender": enquiry.gender,
                    "parent_name": enquiry.parent_name,
                    "parent_mobile_no": enquiry.parent_phone,
                    "address": enquiry.address,
                    "email": enquiry.email,
                    "course_type": enquiry.course_type.id,  # Use .id for <select> fields
                    "course": enquiry.course.id,
                    "pu_percentage": enquiry.percentage_12th, 
                }
            })
        else:
            return JsonResponse({"error": "Provided enquiry is not for Degree form"}, status=404)
    except Enquiry1.DoesNotExist:
        return JsonResponse({"error": "Enquiry number not found"}, status=404)


from django.shortcuts import render

@custom_login_required
def enquiry_print_form(request):
    # You can use any template path; adjust if needed
    return render(request, 'admission/enquiry_print_form.html')

#export all payments getting total paid and pending fee

from django.http import HttpResponse

import openpyxl

from .models import Student

@custom_login_required 
def export_payments_excel(request):

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Student Payments"
 
    # Header Row

    ws.append([

        'Payment Date',

        'Receipt No',

        'Receipt Date',

        'Branch Code',

        'Admission No',

        'Name',

        'Course',
 
        'Tuition Fee',

        'Tuition Fee Paid',

        'Tuition Pending Fee',

        'Tuition Amount Paid',

        'Tuition Due Date',
 
        'Transport Fee',

        'Transport Fee Paid',

        'Transport Pending Fee',

        'Transport Amount Paid',

        'Transport Due Date',
 
        'Hostel Fee',

        'Hostel Fee Paid',

        'Hostel Pending Fee',

        'Hostel Amount Paid',

        'Hostel Due Date',
 
        'Books Fee',

        'Books Fee Paid',

        'Books Pending Fee',

        'Books Amount Paid',

        'Books Due Date',
 
        'Uniform Fee',

        'Uniform Fee Paid',

        'Uniform Pending Fee',

        'Uniform Amount Paid',

        'Uniform Due Date',
 
        'Other Fee',

        'Other Amount Paid',
 
        'Scholarship',

        'Tuition Advance',

        'Final Fee After Advance',
 
        'Total Fee Paid',

        'Total Fee Pending',
 
        'Payment Method',

    ])
 
    students = Student.objects.all().order_by('-id')
 
    for s in students:

        total_fee_paid = sum([

            s.tuition_fee_paid or 0,

            s.transport_fee_paid or 0,

            s.hostel_fee_paid or 0,

            s.books_fee_paid or 0,

            s.uniform_fee_paid or 0,

        ])
 
        total_pending_fee = sum([

            s.tuition_pending_fee or 0,

            s.transport_pending_fee or 0,

            s.hostel_pending_fee or 0,

            s.books_pending_fee or 0,

            s.uniform_pending_fee or 0,

        ])
 
        ws.append([

            s.payment_date.strftime('%Y-%m-%d') if s.payment_date else '',

            s.receipt_no or '',

            s.receipt_date.strftime('%Y-%m-%d') if s.receipt_date else '',

            s.branch_code or '',

            s.admission_no,

            s.name,

            s.course,
 
            s.tuition_fee,

            s.tuition_fee_paid,

            s.tuition_pending_fee,

            s.tuition_amount,

            s.tuition_due_date.strftime('%Y-%m-%d') if s.tuition_due_date else '',
 
            s.transport_fee,

            s.transport_fee_paid,

            s.transport_pending_fee,

            s.transport_amount,

            s.transport_due_date.strftime('%Y-%m-%d') if s.transport_due_date else '',
 
            s.hostel_fee,

            s.hostel_fee_paid,

            s.hostel_pending_fee,

            s.hostel_amount,

            s.hostel_due_date.strftime('%Y-%m-%d') if s.hostel_due_date else '',
 
            s.books_fee,

            s.books_fee_paid,

            s.books_pending_fee,

            s.books_amount,

            s.books_due_date.strftime('%Y-%m-%d') if s.books_due_date else '',
 
            s.uniform_fee,

            s.uniform_fee_paid,

            s.uniform_pending_fee,

            s.uniform_amount,

            s.uniform_due_date.strftime('%Y-%m-%d') if s.uniform_due_date else '',
 
            s.other_fee or '',

            s.other_amount or 0,
 
            s.scholarship or 0,

            s.tuition_advance_amount or 0,

            s.final_fee_after_advance or 0,
 
            total_fee_paid,

            total_pending_fee,
 
            s.payment_method or '',

        ])
 
    response = HttpResponse(

        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

    response['Content-Disposition'] = 'attachment; filename=student_payment_history.xlsx'

    wb.save(response)

    return response

 
#export all payments getting total paid and pending fee
from django.http import HttpResponse

import openpyxl

from django.http import HttpResponse
import openpyxl
from .models import StudentPaymentHistory  # Ensure this model matches your DB

@custom_login_required
def export_payments_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment History"

    # Header Row - Includes all required fields
    ws.append([
        'Payment Date',
        'Receipt No',
        'Receipt Date',
        'Branch Code',
        'Admission No',
        'Name',
        'Course',

        # Tuition Fee
        'Tuition Fee',
        'Tuition Fee Paid',
        'Tuition Pending Fee',
        'Tuition Amount Paid',
        'Tuition Due Date',

        # Transport
        'Transport Fee',
        'Transport Fee Paid',
        'Transport Pending Fee',
        'Transport Amount Paid',
        'Transport Due Date',

        # Hostel
        'Hostel Fee',
        'Hostel Fee Paid',
        'Hostel Pending Fee',
        'Hostel Amount Paid',
        'Hostel Due Date',

        # Books
        'Books Fee',
        'Books Fee Paid',
        'Books Pending Fee',
        'Books Amount Paid',
        'Books Due Date',

        # Uniform
        'Uniform Fee',
        'Uniform Fee Paid',
        'Uniform Pending Fee',
        'Uniform Amount Paid',
        'Uniform Due Date',

        # Other
        'Other Fee',
        'Other Amount Paid',
        'Other Due Date',

        # Scholarship / Advance
        'Scholarship',
        'Tuition Advance',
        'Final Fee After Advance',

        # Totals
        'Total Fee Paid',
        'Total Fee Pending',

        # Payment Meta
        'Payment Method',
    ])

    payments = StudentPaymentHistory.objects.all().order_by('-id')

    for p in payments:
        total_fee_paid = sum([
            p.tuition_fee_paid or 0,
            p.transport_fee_paid or 0,
            p.hostel_fee_paid or 0,
            p.books_fee_paid or 0,
            p.uniform_fee_paid or 0,
        ])

        total_pending_fee = sum([
            p.tuition_pending_fee or 0,
            p.transport_pending_fee or 0,
            p.hostel_pending_fee or 0,
            p.books_pending_fee or 0,
            p.uniform_pending_fee or 0,
        ])

        ws.append([
            p.payment_date.strftime('%Y-%m-%d') if p.payment_date else '',
            p.receipt_no or '',
            p.receipt_date.strftime('%Y-%m-%d') if p.receipt_date else '',
            p.branch_code or '',
            p.admission_no,
            p.name,
            p.course,

            p.tuition_fee,
            p.tuition_fee_paid,
            p.tuition_pending_fee,
            p.tuition_amount,
            p.tuition_due_date.strftime('%Y-%m-%d') if p.tuition_due_date else '',

            p.transport_fee,
            p.transport_fee_paid,
            p.transport_pending_fee,
            p.transport_amount,
            p.transport_due_date.strftime('%Y-%m-%d') if p.transport_due_date else '',

            p.hostel_fee,
            p.hostel_fee_paid,
            p.hostel_pending_fee,
            p.hostel_amount,
            p.hostel_due_date.strftime('%Y-%m-%d') if p.hostel_due_date else '',

            p.books_fee,
            p.books_fee_paid,
            p.books_pending_fee,
            p.books_amount,
            p.books_due_date.strftime('%Y-%m-%d') if p.books_due_date else '',

            p.uniform_fee,
            p.uniform_fee_paid,
            p.uniform_pending_fee,
            p.uniform_amount,
            p.uniform_due_date.strftime('%Y-%m-%d') if p.uniform_due_date else '',

            p.other_fee or '',
            p.other_amount or 0,

            p.scholarship or 0,
            p.tuition_advance_amount or 0,
            p.final_fee_after_advance,

            total_fee_paid,
            total_pending_fee,

            p.payment_method or '',
        ])

    # Return Excel as downloadable response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=student_payment_history.xlsx'
    wb.save(response)
    return response


from django.shortcuts import render
from django.utils import timezone
from datetime import timedelta
from admission.models import Enquiry1, PUAdmission, DegreeAdmission

@custom_login_required
def admission_dashboard(request):
    now = timezone.now()
    today = now.date()
    filter_type = request.GET.get('filter', 'all')  # Default to 'all'

    # Determine date range
    if filter_type == 'day':
        date_start = today
        date_end = today
    elif filter_type == 'week':
        date_start = today - timedelta(days=today.weekday())  # Monday
        date_end = date_start + timedelta(days=6)  # Sunday
    elif filter_type == 'month':
        date_start = today.replace(day=1)
        date_end = today
    else:
        date_start = None
        date_end = None

    # Filter admissions by admission_date
    if date_start and date_end:
        pu_admissions = PUAdmission.objects.filter(admission_date__range=(date_start, date_end))
        degree_admissions = DegreeAdmission.objects.filter(admission_date__range=(date_start, date_end))
        enquiries = Enquiry1.objects.filter(enquiry_date__range=(date_start, date_end))
    else:
        pu_admissions = PUAdmission.objects.all()
        degree_admissions = DegreeAdmission.objects.all()
        enquiries = Enquiry1.objects.all()

    # Counts
    total_enquiries = enquiries.count()
    total_pu_admissions = pu_admissions.count()
    total_degree_admissions = degree_admissions.count()

    pu_converted_enquiries = pu_admissions.exclude(enquiry_no__isnull=True).exclude(enquiry_no='').count()
    degree_converted_enquiries = degree_admissions.exclude(enquiry_no__isnull=True).exclude(enquiry_no='').count()

    pu_confirmed = pu_admissions.filter(status__iexact='Confirmed').count()
    pu_pending = pu_admissions.filter(status__iexact='Pending').count()

    degree_confirmed = degree_admissions.filter(status__iexact='Confirmed').count()
    degree_pending = degree_admissions.filter(status__iexact='Pending').count()

    total_confirmed = pu_confirmed + degree_confirmed
    total_pending = pu_pending + degree_pending

    # Conversion rate
    total_converted = pu_converted_enquiries + degree_converted_enquiries
    conversion_rate = (total_converted / total_enquiries) * 100 if total_enquiries else 0

    context = {
        'total_enquiries': total_enquiries,
        'total_pu_admissions': total_pu_admissions,
        'total_degree_admissions': total_degree_admissions,
        'pu_converted_enquiries': pu_converted_enquiries,
        'degree_converted_enquiries': degree_converted_enquiries,
        'confirmed_admissions': total_confirmed,
        'pending_review': total_pending,
        'conversion_rate': round(conversion_rate, 2),
        'active_filter': filter_type,  # for frontend highlighting
    }

    return render(request, 'admission/admission_dashboard.html', context)

 
from django.shortcuts import render, redirect, get_object_or_404
from .models import PUAdmission, DegreeAdmission, ConfirmedAdmission
from django.contrib import messages
from core.utils import log_activity, get_logged_in_user 

@custom_login_required
def pending_admissions(request):
    if request.method == 'POST':
        admission_no = request.POST.get('admission_no')
        action = request.POST.get('action')

        admission = None
        source = None

        degree_adm = DegreeAdmission.objects.filter(admission_no=admission_no).first()
        if degree_adm:
            admission = degree_adm
            source = "DegreeAdmission"
        else:
            pu_adm = PUAdmission.objects.filter(admission_no=admission_no).first()
            if pu_adm:
                admission = pu_adm
                source = "PUAdmission"

        if not admission:
            messages.error(request, f"❌ Admission number {admission_no} not found.")
            return redirect('pending_admissions')

        user = get_logged_in_user(request)  # ✅ get the logged in user

        if action == 'confirm':
            admission.status = 'Confirmed'
            admission.save()
            start_year = admission.admission_date.year
            end_year = start_year + (2 if source == "PUAdmission" else 3)
            academic_year_str = f"{start_year}-{str(end_year)[-2:]}"


# Get or create the academic year object using `year` field
            academic_year_obj, _ = AcademicYear.objects.get_or_create(year=academic_year_str)

            exists = ConfirmedAdmission.objects.filter(
                pu_admission=admission if source == "PUAdmission" else None,
                degree_admission=admission if source == "DegreeAdmission" else None
            ).exists()

            if not exists:
                ConfirmedAdmission.objects.create(
                    pu_admission=admission if source == "PUAdmission" else None,
                    degree_admission=admission if source == "DegreeAdmission" else None,
                    student_name=admission.student_name,
                    course=str(admission.course),
                    admission_date=admission.admission_date,
                    documents_complete=admission.document_submitted,
                    tuition_advance_amount=getattr(admission, "tuition_advance_amount", None),
                    academic_year=academic_year_obj,
                    current_year=1 if source == "PUAdmission" else None,
                    semester=1 if source == "DegreeAdmission" else None,
                )

            messages.success(request, f"✅ Admission {admission_no} confirmed successfully.")

            # ✅ Log activity
            log_activity(user, 'confirmed admission', admission)

        elif action == 'review':
            admission.status = 'Review'
            admission.save()
            messages.warning(request, f"🔍 Admission {admission_no} marked for review.")

            # ✅ Log activity
            log_activity(user, 'marked admission for review', admission)

        elif action == 'reject':
            admission.status = 'Rejected'
            admission.save()
            messages.error(request, f"❌ Admission {admission_no} rejected.")

            # ✅ Log activity
            log_activity(user, 'rejected admission', admission)

        return redirect('pending_admissions')

    # Fetch PU & Degree pending admissions
    pu_pending = PUAdmission.objects.filter(status__in=['Pending', 'Review'])
    degree_pending = DegreeAdmission.objects.filter(status__in=['Pending', 'Review'])

    # ✅ Map function
    # @custom_login_required
    def admission_dict(adm, program_type):
        return {
            "admission_no": adm.admission_no,
            "student_name": adm.student_name,
            "course": str(adm.course),
            "admission_date": adm.admission_date,
            "documents_complete": adm.document_submitted,
            "status": adm.status,
            # "tuition_advance_amount": getattr(adm, "tuition_advance_amount", 0) or 0,
            "application_fee": getattr(adm, "application_fee", 0) or 0, 
            "pk": adm.pk,
            "course_type": program_type,
        }

    admissions_list = (
        [admission_dict(adm, "PU") for adm in pu_pending] +
        [admission_dict(adm, "Degree") for adm in degree_pending]
    )

    return render(request, "admission/pending_admissions.html", {
        "admissions_list": admissions_list
    })
 
 
from django.shortcuts import redirect
from django.db import transaction
from django.contrib import messages
from master.models import StudentDatabase
from .models import ConfirmedAdmission, PUAdmission, DegreeAdmission
from .utils import generate_student_credentials  # Ensure this exists
from .utils import generate_parent_credentials  # Ensure this exists

@custom_login_required
def generate_student_userid(request, admission_no):
    user = get_logged_in_user(request)  # ✅ Your custom auth logic

    # ------------------ Find ConfirmedAdmission ------------------
    confirmed = (
        ConfirmedAdmission.objects.select_related('pu_admission', 'degree_admission')
        .filter(status='confirmed', pu_admission__admission_no=admission_no)
        .first()
    )

    if not confirmed:
        confirmed = (
            ConfirmedAdmission.objects.select_related('pu_admission', 'degree_admission')
            .filter(status='confirmed', degree_admission__admission_no=admission_no)
            .first()
        )

    if not confirmed:
        messages.error(request, f"No confirmed admission found for Admission No: {admission_no}")
        return redirect('confirmed_admissions')

    admission = confirmed.pu_admission or confirmed.degree_admission
    is_degree = isinstance(admission, DegreeAdmission)

    try:
        with transaction.atomic():

            # ------------------ Student ID Generation ------------------
            if not confirmed.student_userid:
                existing_student_ids = set(
                    ConfirmedAdmission.objects.exclude(student_userid__isnull=True)
                    .values_list('student_userid', flat=True)
                )
                student_userid, student_password = generate_student_credentials(existing_student_ids)
                confirmed.student_userid = student_userid
                confirmed.student_password = student_password

            # ------------------ Parent ID Generation ------------------
            if not confirmed.parent_userid:
                existing_parent_ids = set(
                    ConfirmedAdmission.objects.exclude(parent_userid__isnull=True)
                    .values_list('parent_userid', flat=True)
                )
                parent_userid, parent_password = generate_parent_credentials(existing_parent_ids)
                confirmed.parent_userid = parent_userid
                confirmed.parent_password = parent_password

            # ------------------ Pick Parent Contact ------------------
            parent_name, parent_email, parent_phone = None, None, None

            if admission.primary_guardian == "father":
                parent_name = getattr(admission, "father_name", None)
                parent_email = getattr(admission, "father_email", None)
                parent_phone = getattr(admission, "father_mobile_no", None)
            elif admission.primary_guardian == "mother":
                parent_name = getattr(admission, "mother_name", None)
                parent_email = getattr(admission, "mother_email", None)
                parent_phone = getattr(admission, "mother_phone_no", None)
            elif admission.primary_guardian == "guardian":
                parent_name = getattr(admission, "guardian_name", None)
                parent_email = getattr(admission, "guardian_email", None)
                parent_phone = getattr(admission, "guardian_phone_no", None)

            # ------------------ Student Database Sync ------------------
            StudentDatabase.objects.update_or_create(
                pu_admission=admission if not is_degree else None,
                degree_admission=admission if is_degree else None,
                defaults={
                    'student_name': admission.student_name,
                    'course': admission.course if admission.course else None,
                    'course_type': getattr(admission, 'course_type', None),
                    'quota_type': getattr(admission, 'quota_type', None) if is_degree else None,
                    'student_userid': confirmed.student_userid,
                    'student_phone_no': getattr(admission, 'student_phone_no', None),
                    'father_name': getattr(admission, 'father_name', None),
                    'academic_year': confirmed.academic_year.year if confirmed.academic_year else None,
                    'current_year': confirmed.current_year,
                    'semester': confirmed.semester,
                   
                }
            )

            # ------------------ Save Everything ------------------
            confirmed.save()
            log_activity(user, 'generated', confirmed)

            messages.success(
                request,
                f"✅ Student & Parent IDs generated and saved for Admission No: {admission_no}"
            )

    except Exception as e:
        # Rollback IDs if database update fails
        print(f"[ERROR] StudentDatabase save failed for {admission_no}: {e}")
        confirmed.student_userid = None
        confirmed.student_password = None
        confirmed.parent_userid = None
        confirmed.parent_password = None
        confirmed.save(update_fields=[
            "student_userid", "student_password",
            "parent_userid", "parent_password"
        ])
        messages.error(
            request,
            f"⚠️ IDs were generated but not saved for Admission No: {admission_no}. Please try again."
        )

    return redirect('confirmed_admissions')

 
 
 
from .models import PUAdmission, DegreeAdmission, ConfirmedAdmission
 
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .utils import generate_student_credentials  # Assuming you have a utility function

@custom_login_required
def confirmed_admissions(request):
    if request.method == 'POST':
        admission_id = request.POST.get('admission_id')
        admission = get_object_or_404(ConfirmedAdmission, id=admission_id)

        if not admission.student_userid:
            userid, password = generate_student_credentials()
            admission.student_userid = userid
            admission.student_password = password
            admission.save()

            messages.success(
                request,
                f"Student credentials generated for {admission.student_name} (ID: {userid})"
            )
        else:
            messages.info(
                request,
                f"Credentials already exist for {admission.student_name} (ID: {admission.student_userid})"
            )

        return redirect('confirmed_admissions')

    admissions_list = ConfirmedAdmission.objects.select_related('pu_admission', 'degree_admission').all()

    return render(request, "admission/confirmed_admissions.html", {
        "admissions_list": admissions_list
    })
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.conf import settings
from .models import ConfirmedAdmission

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.conf import settings
from .models import ConfirmedAdmission

def send_admission_email(request, admission_id):
    if request.method == "POST":
        admission = get_object_or_404(ConfirmedAdmission, id=admission_id)
        pu_adm = getattr(admission, "pu_admission", None)
        degree_adm = getattr(admission, "degree_admission", None)

        if pu_adm:
            student_name = pu_adm.student_name
            course = pu_adm.course.name if pu_adm.course else ""
            year = pu_adm.admission_date.year
            student_userid = admission.student_userid
            student_password = admission.student_password
            parent_userid = admission.parent_userid
            parent_password = admission.parent_password
            parent_name = pu_adm.parent_name()
            student_email = pu_adm.student_email() if callable(pu_adm.student_email) else pu_adm.student_email
            parent_email = pu_adm.parent_email() if callable(pu_adm.parent_email) else pu_adm.parent_email

        elif degree_adm:
            student_name = degree_adm.student_name
            course = degree_adm.course.name if degree_adm.course else ""
            year = degree_adm.admission_date.year
            student_userid = admission.student_userid
            student_password = admission.student_password
            parent_userid = admission.parent_userid
            parent_password = admission.parent_password
            parent_name = degree_adm.parent_name()
            student_email = degree_adm.student_email() if callable(degree_adm.student_email) else degree_adm.student_email
            parent_email = degree_adm.parent_email() if callable(degree_adm.parent_email) else degree_adm.parent_email

        try:
            if student_email:
                student_message = render_to_string(
                    'admission/admission_email_student.html',
                    {
                        'student_name': student_name,
                        'student_userid': student_userid,
                        'student_password': student_password,
                        'course': course,
                        'year': year,
                    }
                )
                email = EmailMessage(
                    subject="Admission Confirmation - Student",
                    body=student_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[student_email],
                )
                email.content_subtype = "html"
                email.send()

            if parent_email:
                parent_message = render_to_string(
                    'admission/admission_email_parent.html',
                    {
                        'parent_name': parent_name,
                        'student_name': student_name,
                        'parent_userid': parent_userid,
                        'parent_password': parent_password,
                        'course': course,
                        'year': year,
                    }
                )
                email = EmailMessage(
                    subject="Admission Confirmation - Parent",
                    body=parent_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[parent_email],
                )
                email.content_subtype = "html"
                email.send()

            messages.success(request, "Emails sent successfully.")

        except Exception as e:
            messages.error(request, f"Error sending email: {e}")

        # ✅ Redirect back and mark which admission was just sent
        referer = request.META.get("HTTP_REFERER", "/")
        return redirect(f"{referer}?sent={admission_id}")

    return redirect('/')

 
 
from django.shortcuts import get_object_or_404, redirect
from .models import DegreeAdmission

@custom_login_required 
def update_status(request, pk, new_status):
    admission = get_object_or_404(DegreeAdmission, pk=pk)
    admission.status = new_status
    admission.save()
    # Redirect as appropriate, e.g., to pending or confirmed list
    return redirect('pending_admissions')




from django.shortcuts import redirect, get_object_or_404
from .models import Enquiry1, Enquiry2
from  core.utils import get_logged_in_user, log_activity

@custom_login_required
def convert_enquiry(request, enquiry_no):
    user = get_logged_in_user(request)

    if enquiry_no.startswith("PU-ENQ-"):
        enquiry = get_object_or_404(Enquiry1, enquiry_no=enquiry_no)
        log_activity(user, 'converted', enquiry)
        return redirect('admission_form', enquiry_no=enquiry.enquiry_no)
    elif enquiry_no.startswith("DEG-ENQ-"):
        enquiry = get_object_or_404(Enquiry2, enquiry_no=enquiry_no)
        log_activity(user, 'converted', enquiry)
        return redirect('degree_admission_form', enquiry_no=enquiry.enquiry_no)
    else:
        log_activity(user, 'convert_failed', enquiry_no)
        return redirect('enquiry_list')

from django.shortcuts import render
from .models import PUAdmission, DegreeAdmission

@custom_login_required
def reports(request):
    # Course-wise admission counts
    pu_count = PUAdmission.objects.count()
    degree_count = DegreeAdmission.objects.count()

    # Statistics (summing both models)
    total_admissions = pu_count + degree_count
    confirmed_admissions = PUAdmission.objects.filter(status='Confirmed').count() + DegreeAdmission.objects.filter(status='Confirmed').count()
    pending_admissions = PUAdmission.objects.filter(status='Pending').count() + DegreeAdmission.objects.filter(status='Pending').count()
    rejected_admissions = PUAdmission.objects.filter(status='Rejected').count() + DegreeAdmission.objects.filter(status='Rejected').count()
    review_admissions = PUAdmission.objects.filter(status='Review').count() + DegreeAdmission.objects.filter(status='Review').count()

    context = {
        "pu_count": pu_count,
        "degree_count": degree_count,
        "total_admissions": total_admissions,
        "confirmed_admissions": confirmed_admissions,
        "pending_admissions": pending_admissions,
        "rejected_admissions": rejected_admissions,
        "review_admissions": review_admissions,
    }
    return render(request, "admission/reports.html", context)



from django.db.models import Value, CharField

from admission.models import Enquiry1, Enquiry2, PUAdmission, DegreeAdmission

from master.models import UserCustom
from django.db.models import Value, CharField
from django.shortcuts import render
from admission.models import Enquiry1, Enquiry2, PUAdmission, DegreeAdmission

@custom_login_required
def converted_enquiry_list(request):
    # Get enquiry numbers that are converted
    pu_admissions = PUAdmission.objects.exclude(enquiry_no__isnull=True).exclude(enquiry_no='').values('enquiry_no', 'admission_no')
    deg_admissions = DegreeAdmission.objects.exclude(enquiry_no__isnull=True).exclude(enquiry_no='').values('enquiry_no', 'admission_no')

    # Create lookup dictionaries for admission numbers
    pu_adm_dict = {entry['enquiry_no']: entry['admission_no'] for entry in pu_admissions}
    deg_adm_dict = {entry['enquiry_no']: entry['admission_no'] for entry in deg_admissions}

    pu_enquiry_nos = list(pu_adm_dict.keys())
    deg_enquiry_nos = list(deg_adm_dict.keys())

    # Get PU and DEG enquiries and annotate them
    enquiries1 = Enquiry1.objects.filter(enquiry_no__in=pu_enquiry_nos).annotate(enquiry_type=Value('PU', output_field=CharField()))
    enquiries2 = Enquiry2.objects.filter(enquiry_no__in=deg_enquiry_nos).annotate(enquiry_type=Value('DEG', output_field=CharField()))

    enquiries = list(enquiries1) + list(enquiries2)

    # Add created_by usernames
    user_ids = [e.created_by for e in enquiries if e.created_by]
    users = UserCustom.objects.filter(id__in=user_ids).values('id', 'username')
    user_dict = {user['id']: user['username'] for user in users}

    for enquiry in enquiries:
        enquiry.created_by_username = user_dict.get(enquiry.created_by, "Unknown")
        # Add admission number
        if enquiry.enquiry_type == 'PU':
            enquiry.admission_no = pu_adm_dict.get(enquiry.enquiry_no, 'N/A')
        elif enquiry.enquiry_type == 'DEG':
            enquiry.admission_no = deg_adm_dict.get(enquiry.enquiry_no, 'N/A')

    return render(request, 'admission/converted_enquiry_list.html', {
        'enquiries': enquiries,
    })

# def confirmed_admissions(request):
#     if request.method == 'POST':
#         admission_id = request.POST.get('admission_id')
#         admission = get_object_or_404(ConfirmedAdmission, id=admission_id)
#         if not admission.student_userid:
#             userid, password = generate_student_credentials()
#             admission.student_userid = userid
#             admission.student_password = password
#             admission.save()
#         return redirect('confirmed_admissions')

#     admissions_list = ConfirmedAdmission.objects.all()
#     return render(request, "admission/confirmed_admissions.html", {"admissions_list": admissions_list})

#fee dashboard


from decimal import Decimal
from django.db.models import F, Sum, DecimalField, ExpressionWrapper, Q
from django.db.models.functions import Coalesce
from django.shortcuts import render
from .models import Student, StudentPaymentHistory
from django.db.models import DecimalField, ExpressionWrapper, F, Sum
from django.db.models.functions import Coalesce


@custom_login_required
def to_decimal(value):
    try:
        return Decimal(value or 0)
    except:
        return Decimal(0)

@custom_login_required
def dashboard_view(request):
    total_declared_fee = Decimal(0)
    total_advance_fee = Decimal(0)

    # Loop through each student and compute declared fee minus advance
    for student in Student.objects.all():
        tuition_fee = to_decimal(student.tuition_fee)
        books_fee = to_decimal(student.books_fee)
        uniform_fee = to_decimal(student.uniform_fee)
        transport_fee = to_decimal(student.transport_fee)
        hostel_fee = to_decimal(student.hostel_fee)
        other_fee = to_decimal(student.other_amount)
        advance = to_decimal(student.tuition_advance_amount)

        declared_fee = tuition_fee + books_fee + uniform_fee + transport_fee + hostel_fee + other_fee 
        final_fee = declared_fee 

        total_declared_fee += final_fee
        total_advance_fee += advance

    # Total collected fee
    total_collected_fee = Student.objects.aggregate(
        total=Sum(
            ExpressionWrapper(
                Coalesce(F('tuition_fee_paid'), 0) +
                Coalesce(F('transport_fee_paid'), 0) +
                Coalesce(F('hostel_fee_paid'), 0) +
                Coalesce(F('books_fee_paid'), 0) +
                Coalesce(F('uniform_fee_paid'), 0) +
                Coalesce(F('other_amount'), 0) +
                Coalesce(F('tuition_advance_amount'), 0),
                output_field=DecimalField()
            )
        )
    )['total'] or 0

    total_pending_fee = total_declared_fee - total_collected_fee

    collected_fee_percentage = round(
        (total_collected_fee / total_declared_fee) * 100
    ) if total_declared_fee else 0

    remaining_fee_percentage = round(
        (total_pending_fee / total_declared_fee) * 100
    ) if total_declared_fee else 0

    # Search functionality
    search_query = request.GET.get('search', '').strip()
    matching_students = []
    history = []

    if search_query:
        matching_students = Student.objects.filter(
            Q(admission_no__icontains=search_query) |
            Q(name__icontains=search_query)
        )

        history = StudentPaymentHistory.objects.filter(
            Q(admission_no__icontains=search_query) |
            Q(name__icontains=search_query)
        ).annotate(
            base_fee=Coalesce(F('tuition_fee'), Decimal(0)),
            fee_paid=Coalesce(F('tuition_fee_paid'), Decimal(0)) + Coalesce(F('scholarship'), Decimal(0)),
            final_fee=Coalesce(F('final_fee_after_advance'), Decimal(0))
        ).order_by('-payment_date')

    student_data = Student.objects.all()

    context = {
        'total_declared_fee': total_declared_fee,
        'total_collected_fee': total_collected_fee,
        'total_pending_fee': total_pending_fee,
        'total_advance_fee': total_advance_fee,
        'collected_fee_percentage': collected_fee_percentage,
        'remaining_fee_percentage': remaining_fee_percentage,
        'search_query': search_query,
        'history': history,
        'student_data': student_data,
        'matching_students': matching_students,
    }

    return render(request, 'admission/student_fee_dashboard.html', context)
from django.http import HttpResponse
from django.template.loader import render_to_string
from io import BytesIO
from .models import DegreeAdmission, PUAdmission
from django.utils import timezone

def render_to_pdf(template_src, context_dict={}):
    """
    Render a Django template to PDF using xhtml2pdf
    """
    from xhtml2pdf import pisa

    template = render_to_string(template_src, context_dict)
    result = BytesIO()
    pdf = pisa.CreatePDF(src=template, dest=result)
    if not pdf.err:
        return result.getvalue()
    return None


@custom_login_required
def download_degree_admission_fee_receipt(request, admission_no):
    admission = DegreeAdmission.objects.filter(admission_no=admission_no).first()
    if not admission:
        return HttpResponse("Admission not found.", status=404)

    context = {
        'admission': admission,
        'date_now': timezone.now(),
    }
    pdf_file = render_to_pdf('admission/degree_admission_form_fee_receipt.html', context)

    if not pdf_file:
        return HttpResponse("Error generating PDF", status=500)

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=degree_fee_receipt_{admission_no}.pdf'
    return response


@custom_login_required
def download_pu_admission_fee_receipt(request, admission_no):
    admission = PUAdmission.objects.filter(admission_no=admission_no).first()
    if not admission:
        return HttpResponse("Admission not found.", status=404)

    context = {
        'admission': admission,
        'date_now': timezone.now(),
    }
    pdf_file = render_to_pdf('admission/admission_form_fee_receipt.html', context)

    if not pdf_file:
        return HttpResponse("Error generating PDF", status=500)

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=pu_fee_receipt_{admission_no}.pdf'
    return response






from django.shortcuts import render
from django.utils import timezone
from datetime import timedelta
from django.db.models import Sum, F, DecimalField, ExpressionWrapper, Case, When, Value
from django.db.models.functions import Coalesce
from .models import Student

@custom_login_required
def fee_management(request):
    now = timezone.now()
    today = now.date()
    filter_type = request.GET.get('filter', 'all')  # Default to 'all'

    # Determine date range
    if filter_type == 'day':
        date_start = today
        date_end = today
    elif filter_type == 'week':
        date_start = today - timedelta(days=today.weekday())  # Monday
        date_end = date_start + timedelta(days=6)  # Sunday
    elif filter_type == 'month':
        date_start = today.replace(day=1)
        date_end = today
    else:
        date_start = None
        date_end = None

    # Apply date filter
    if date_start and date_end:
        students = Student.objects.filter(receipt_date__range=(date_start, date_end))
    else:
        students = Student.objects.all()

    decimal_type = DecimalField(max_digits=12, decimal_places=2)

    # Define deduction logic
    deduction_expr = Case(
        When(course__icontains="PUC", then=Value(5000)),
        default=Value(10000),
        output_field=decimal_type
    )

    # Total Declared Fee
    total_declared_fee = students.aggregate(
        total=Sum(
            ExpressionWrapper(
                Coalesce(F('tuition_fee'), 0) +
                Coalesce(F('transport_amount'), 0) +
                Coalesce(F('hostel_amount'), 0) +
                Coalesce(F('books_fee'), 0) +
                Coalesce(F('uniform_fee'), 0) +
                Coalesce(F('other_amount'), 0) - deduction_expr,
                output_field=decimal_type
            )
        )
    )['total'] or 0

    # Total Collected Fee
    total_collected_fee = students.aggregate(
        total=Sum(
            ExpressionWrapper(
                Coalesce(F('tuition_fee_paid'), 0) +
                Coalesce(F('transport_fee_paid'), 0) +
                Coalesce(F('hostel_fee_paid'), 0) +
                Coalesce(F('books_fee_paid'), 0) +
                Coalesce(F('uniform_fee_paid'), 0) +
                Coalesce(F('other_amount'), 0),
                output_field=decimal_type
            )
        )
    )['total'] or 0

    # Pending & Percentages
    total_pending_fee = total_declared_fee - total_collected_fee
    collected_fee_percentage = round((total_collected_fee / total_declared_fee) * 100) if total_declared_fee else 0
    remaining_fee_percentage = round((total_pending_fee / total_declared_fee) * 100) if total_declared_fee else 0

    context = {
        'total_declared_fee': total_declared_fee,
        'total_collected_fee': total_collected_fee,
        'total_pending_fee': total_pending_fee,
        'collected_fee_percentage': collected_fee_percentage,
        'remaining_fee_percentage': remaining_fee_percentage,
        'active_filter': filter_type,  # for frontend highlighting
    }

    return render(request, 'admission/fee_management.html', context)    




from django.http import JsonResponse
from master.models import Course

@custom_login_required
def get_courses_by_type(request):
    course_type_id = request.GET.get('course_type_id')
    courses = Course.objects.filter(course_type_id=course_type_id).values('id', 'name')
    return JsonResponse(list(courses), safe=False)

from django.http import JsonResponse
from django.shortcuts import render
from master.models import AcademicYear, CourseType, Course, StudentDatabase

@custom_login_required
def get_course_types_by_academic(request):
    academic_year_id = request.POST.get('academic_year_id')
    course_types = CourseType.objects.filter(academic_year_id=academic_year_id)
    data = {
        "course_types": [
            {"id": ct.id, "name": ct.name}
            for ct in course_types
        ]
    }
    return JsonResponse(data)


@custom_login_required
# AJAX: Get Courses for selected CourseType
def get_courses_by_type(request):
    course_type_id = request.POST.get('course_type_id')
    courses = Course.objects.filter(course_type_id=course_type_id).values('id', 'name')
    return JsonResponse({'courses': list(courses)})


from django.http import JsonResponse
from master.models import Course,Semester

# AJAX: Get Courses for selected CourseType
@custom_login_required
def get_courses_by_type(request):
    course_type_id = request.POST.get('course_type_id')

    if not course_type_id:
        return JsonResponse({'error': 'Course type ID is required'}, status=400)

    try:
        course_type_id = int(course_type_id)
    except ValueError:
        return JsonResponse({'error': 'Invalid course type ID format'}, status=400)

    courses = Course.objects.filter(course_type_id=course_type_id).values('id', 'name')
    return JsonResponse({'courses': list(courses)})


# AJAX: Get Semesters by Course
@custom_login_required
def get_sem_by_course(request):
    course_id = request.GET.get("course_id")

    if not course_id:
        return JsonResponse({'error': 'Course ID not provided'}, status=400)

    try:
        course_id = int(course_id)
        course = Course.objects.get(id=course_id)

        semester_list = []

        # Handle course type logic
        course_type_name = course.course_type.name.strip().lower()

        total = (
            course.duration_years if course_type_name == "puc regular"
            else course.total_semesters or 0
        )

        for i in range(1, total + 1):
            semester_list.append({
                'number': i,
                'name': f"{course.name} {i}"
            })

        if not semester_list:
            semester_list.append({'number': 0, 'name': "NOT APPLICABLE"})

        return JsonResponse({'semesters': semester_list})

    except ValueError:
        return JsonResponse({'error': 'Invalid course ID format'}, status=400)

    except Course.DoesNotExist:
        return JsonResponse({'error': 'Invalid course ID'}, status=404)




@custom_login_required
def student_fee_list(request):
    course_type_id = request.GET.get('course_type')
    course_id = request.GET.get('course')
    academic_year_id = request.GET.get('academic_year')
    semester_number = request.GET.get('semester')

    # Handle undefined or empty semester gracefully
    try:
        semester_number = int(semester_number) if semester_number else None
    except (ValueError, TypeError):
        semester_number = None

    # Fetch course types and academic years for dropdowns
    course_types = CourseType.objects.all()
    academic_years = AcademicYear.objects.all()

    # Get the actual academic year string (e.g., "2025–2027") for comparison
    academic_year_obj = AcademicYear.objects.filter(id=academic_year_id).first()
    academic_year_str = academic_year_obj.year if academic_year_obj else None

    # Fetch courses based on course_type_id, if provided
    courses = Course.objects.filter(course_type_id=course_type_id) if course_type_id else Course.objects.all()

    # Fetch semesters based on selected course (to populate semester dropdown)
    semesters = []
    if course_id:
        semesters = Semester.objects.filter(course_id=course_id).order_by('number')

    students_data = []
    fee_names = []

    if all([course_type_id, course_id, academic_year_id, semester_number is not None]):
        course_type = CourseType.objects.filter(id=course_type_id).first()

        if course_type and "PU" in course_type.name.upper():
            students = StudentDatabase.objects.filter(
                course_type_id=course_type_id,
                course_id=course_id,
                current_year=semester_number,
                status="Active",
                student_userid__isnull=False,
                pu_admission__isnull=False,
                academic_year__iexact=academic_year_str  # ✅ safer comparison
            ).select_related('pu_admission', 'degree_admission')
        else:
            students = StudentDatabase.objects.filter(
                course_type_id=course_type_id,
                course_id=course_id,
                semester=semester_number,
                status="Active",
                student_userid__isnull=False,
                degree_admission__isnull=False,
                academic_year__iexact=academic_year_str  # ✅ safer comparison
            ).select_related('pu_admission', 'degree_admission')

        fee_masters = FeeMaster.objects.filter(
            program_type_id=course_type_id,
            combination_id=course_id
        ).order_by('fee_name')

        fee_names = [fm.fee_name for fm in fee_masters] if fee_masters.exists() else ["No fee details available"]

        for student in students:
            admission = student.pu_admission or student.degree_admission
            if not admission:
                continue

            if not student.academic_year:
                print(f"⚠️ Missing academic year for student: {student.student_name}, ID: {student.student_userid}")
                continue

            # ✅ Corrected indentation: append only if academic_year is present
            students_data.append({
                'admission_no': admission.admission_no,
                'student_id': student.student_userid,
                'student_name': student.student_name,
                'admission_course_type': admission.course_type.name if admission.course_type else "",
                'admission_course': admission.course.name if admission.course else "",
                'admission_dob': getattr(admission, 'dob', ""),
                'admission_academic_year': student.academic_year or "",  # ✅ updated
                'admission_father_name': getattr(admission, 'father_name', ""),
                'admission_father_mobile_no': getattr(admission, 'father_mobile_no', ""),
                'fees': {fm.fee_name: fm.fee_amount for fm in fee_masters}
            })
    else:
        print("⚠️ Incomplete parameters. Please ensure all filters are provided.")

    context = {
        'course_types': course_types,
        'courses': courses,
        'academic_years': academic_years,
        'semesters': semesters,  # <-- Pass semesters here to populate dropdown and keep selection
        'selected_academic_year': int(academic_year_id) if academic_year_id else None,
        'selected_course_type': int(course_type_id) if course_type_id else None,
        'selected_course': int(course_id) if course_id else None,
        'selected_semester': semester_number,
        'students': students_data,
        'fee_names': fee_names,
    }

    return render(request, 'admission/student_fee_list.html', context)








from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.contrib import messages
from decimal import Decimal
from admission.models import PUAdmission, DegreeAdmission, StudentFeeCollection
from master.models import FeeMaster, StudentDatabase

@custom_login_required
def fee_collection_collect(request):
    if request.method == 'POST':
        admission_no = request.POST.get('admission_no')
        selected_fee_ids = request.POST.getlist('selected_fees')
        payment_mode = request.POST.get('payment_mode')
        payment_id = request.POST.get('payment_id')

        for fee_id in selected_fee_ids:
            fee_collection = get_object_or_404(StudentFeeCollection, id=fee_id)
            paid_amount_str = request.POST.get(f'paid_amount_{fee_id}', '0')
            paid_amount = Decimal(paid_amount_str) if paid_amount_str else Decimal('0')

            if paid_amount <= 0:
                continue

            new_paid_amount = fee_collection.paid_amount + paid_amount
            new_balance = max(fee_collection.amount - new_paid_amount, Decimal('0'))

            if new_balance == 0:
                status = 'Paid'
            elif new_paid_amount > 0:
                status = 'Partial'
            else:
                status = 'Pending'

            fee_collection.paid_amount = new_paid_amount
            fee_collection.balance_amount = new_balance
            fee_collection.payment_mode = payment_mode
            fee_collection.payment_id = payment_id
            fee_collection.payment_date = timezone.now()
            fee_collection.status = status
            fee_collection.save()

        messages.success(request, "Payment recorded successfully.")
        return redirect('student_fee_list')

    admission_no = request.GET.get('admission_no')
    admission = None
    student_type = None

    try:
        admission = PUAdmission.objects.get(admission_no=admission_no)
        student_type = 'PU'
    except PUAdmission.DoesNotExist:
        try:
            admission = DegreeAdmission.objects.get(admission_no=admission_no)
            student_type = 'DEGREE'
        except DegreeAdmission.DoesNotExist:
            messages.error(request, "Admission not found.")
            return redirect('student_fee_list')

    student_db = StudentDatabase.objects.filter(
        pu_admission=admission if student_type == 'PU' else None,
        degree_admission=admission if student_type == 'DEGREE' else None
    ).first()

    fees_master = FeeMaster.objects.filter(program_type=admission.course_type)
    fee_collections = []

    for fee in fees_master:
        fee_name = fee.fee_name.lower().strip()

        if fee_name in ['tuition fee', 'hostel fee']:
            # For Tuition/Hostel Fee, check if any installments already exist
            existing_installments = StudentFeeCollection.objects.filter(
                admission_no=admission.admission_no,
                fee_type=fee
            ).order_by('installment_number')

            if existing_installments.exists():
                fee_collections.extend(existing_installments)
            else:
                # Create 3 installments
                for inst in range(1, 4):
                    inst_amount = round(Decimal(fee.fee_amount) / 3, 2)
                    new_record = StudentFeeCollection.objects.create(
                        admission_no=admission.admission_no,
                        student_userid=getattr(student_db, 'student_userid', '') if student_db else '',
                        academic_year=admission.academic_year,
                        semester=getattr(admission, 'semester', None),
                        fee_type=fee,
                        installment_number=inst,
                        amount=inst_amount,
                        paid_amount=Decimal('0'),
                        balance_amount=inst_amount,
                        status='Pending'
                    )
                    fee_collections.append(new_record)
        else:
            # For ALL OTHER FEES, use full amount as is
            existing = StudentFeeCollection.objects.filter(
                admission_no=admission.admission_no,
                fee_type=fee,
                installment_number=None
            ).first()

            if existing:
                fee_collections.append(existing)
            else:
                new_record = StudentFeeCollection.objects.create(
                    admission_no=admission.admission_no,
                    student_userid=getattr(student_db, 'student_userid', '') if student_db else '',
                    academic_year=admission.academic_year,
                    semester=getattr(admission, 'semester', None),
                    fee_type=fee,
                    installment_number=None,
                    amount=Decimal(fee.fee_amount),
                    paid_amount=Decimal('0'),
                    balance_amount=Decimal(fee.fee_amount),
                    status='Pending'
                )
                fee_collections.append(new_record)

    context = {
        'student': {
            'admission_no': admission.admission_no,
            'student_name': admission.student_name,
            'admission_course_type': admission.course_type.name,
            'admission_course': admission.course.name,
            'admission_dob': admission.dob,
            'admission_academic_year': admission.academic_year,
            'admission_father_name': admission.father_name,
            'admission_father_mobile_no': admission.father_mobile_no,
            'category': admission.category,
            'roll_number': getattr(student_db, 'student_userid', '') if student_db else '',
        },
        'fee_collections': fee_collections,
        'now': timezone.now()
    }
    return render(request, 'admission/fee_collection_collect.html', context)


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@custom_login_required
@csrf_exempt
def collect_fee_payment_ajax(request):
    if request.method == "POST":
        admission_no = request.POST.get('admission_no')
        selected_fee_ids = request.POST.get('selected_fee_ids', '').split(',')
        payment_mode = request.POST.get('payment_mode')
        amount_paid = request.POST.get('amount_paid')

        try:
            amount_paid = float(amount_paid)
            selected_fee_ids = [int(i) for i in selected_fee_ids if i.isdigit()]
            fees = StudentFeeCollection.objects.filter(id__in=selected_fee_ids, admission_no=admission_no)

            for fee in fees:
                fee.paid_amount += amount_paid
                fee.balance_amount = fee.amount - fee.paid_amount
                if fee.paid_amount >= fee.amount:
                    fee.status = "Paid"
                elif fee.paid_amount > 0:
                    fee.status = "Partial"
                fee.payment_mode = payment_mode
                fee.payment_date = timezone.now().date()
                fee.save()

            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request."})









from django.shortcuts import render, redirect
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from admission.models import StudentFeeCollection, StudentPaymentHistory


@custom_login_required
@csrf_exempt
def collect_fee_payment_page(request):
    now = timezone.now()

    if request.method == "POST":
        admission_no = request.POST.get("admission_no")
        fee_ids = request.POST.getlist("selected_fees")
        payment_mode = request.POST.get("payment_mode")
        payment_reference = request.POST.get("payment_reference", "").strip()

        for fee_id in fee_ids:
            try:
                paid_amount_str = request.POST.get(f"paid_amount_{fee_id}", "0").strip()
                if not paid_amount_str:
                    continue

                paid_amount = float(paid_amount_str)
                fee_obj = StudentFeeCollection.objects.get(id=fee_id)

                # Update fee collection
                fee_obj.paid_amount += paid_amount
                fee_obj.balance_amount = max(fee_obj.amount - fee_obj.paid_amount, 0)
                fee_obj.payment_mode = payment_mode
                fee_obj.payment_reference = payment_reference
                fee_obj.payment_date = now.date()
                fee_obj.status = "Paid" if fee_obj.balance_amount == 0 else "Partial"
                fee_obj.save()

                # Save payment history
                StudentPaymentHistory.objects.create(
                    admission_no=admission_no,
                    fee=fee_obj,
                    paid_amount=paid_amount,
                    payment_mode=payment_mode,
                    payment_reference=payment_reference,
                )

                print(f"✅ Payment done for Fee ID {fee_id}")

            except Exception as e:
                print(f"❌ Error for Fee ID {fee_id}: {e}")

        return redirect("student_fee_list")

    return redirect("student_fee_list")



# admission/views.py

import qrcode
from io import BytesIO
from django.http import HttpResponse
from django.views.decorators.http import require_GET

# admission/views.py

import qrcode
from io import BytesIO
from django.http import HttpResponse
from django.views.decorators.http import require_GET

@custom_login_required
@require_GET
def generate_qr_dynamic(request):
    amount = request.GET.get("amount")
    if not amount:
        return HttpResponse("Amount is required", status=400)

    upi_id = "9483508971@ybl"
    upi_link = f"upi://pay?pa={upi_id}&pn=Your College Name&am={amount}&cu=INR"

    qr = qrcode.make(upi_link)
    buffer = BytesIO()
    qr.save(buffer)
    buffer.seek(0)

    return HttpResponse(buffer.getvalue(), content_type="image/png")



@custom_login_required
def student_fee_history(request, admission_no):
    history = StudentPaymentHistory.objects.filter(admission_no=admission_no).order_by('-payment_date')
    return render(request, "student_fee_history.html", {"history": history})


